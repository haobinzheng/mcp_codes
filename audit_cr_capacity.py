#!/usr/bin/env python3
"""
audit_cr_capacity.py

Audit Core Router (CR) capacity using:
  1. Primary Source of Truth: 'show isis adjacency' on the source router to find
     the authoritative active bundle ('ae') interface(s) peering with the neighbor.
  2. Secondary Verification: 'show interfaces descriptions | match <target>' to discover
     underlying member links (et-*) and descriptions.
  3. Capacity & Utilization Audit: 'show interface <ae_x>' to determine:
     - Capacity from "Speed" (e.g., Speed: 200Gbps, 800Gbps, 1.6Tbps)
     - Live Input and Output traffic rates (bps and pps)
     - Drained / Idle detection: If multiple AE interfaces exist for a neighbor,
       any AE with negligible/drained traffic (< 10 Mbps) is detected and ignored,
       leaving the true active capacity.
  4. Report Export: Supports exporting full structured reports to Excel (.xlsx) and JSON.
"""

import os
import sys
import re
import json
import asyncio
import argparse
import time
from datetime import datetime
from zoneinfo import ZoneInfo

# Audit command definitions provided for CR capacity audit
DEFAULT_AUDIT_COMMANDS = [
    ("cr01.atl101", "bng01.bna103"),
    ("cr01.atl101", "cr01.aus122"),
    ("cr01.atl103", "cr01.iad101"),
    ("cr01.aus122", "cr01.phx103"),
    ("cr01.dfw101", "cr01.aus121"),
    ("cr01.dfw101", "cr02.mci102"),
    ("cr01.dfw101", "cr02.mci102"),
    ("cr01.dfw101", "cr01.mci101"),
    ("cr01.dfw101", "cr01.aus121"),
    ("cr01.iad101", "cr01.rdu103"),
    ("cr01.iad101", "cr01.rdu101"),
    ("cr01.lax101", "cr01.phx102"),
    ("cr01.ord101", "cr01.atl101"),
    ("cr01.sjc101", "cr02.slc102"),
    ("cr01.slc101", "cr01.den103"),
    ("cr02.mci102", "bng01.oma102"),
    ("cr02.mci102", "cr01.ord101"),
    ("cr02.mci102", "cr01.den102"),
]

# Path to gnetch.sh if stubby is unavailable
GNETCH_SH_PATH = "/usr/local/google/home/mikezh/Coding/gfiber/bin/gnetch.sh"

# Idle traffic threshold: rates below 10 Mbps are considered control-plane keepalives only
IDLE_TRAFFIC_THRESHOLD_BPS = 10_000_000  # 10 Mbps

RATE_LIMIT = 0.5  # Seconds rate limit per device
rate_limiters = {}


def get_pacific_timestamp() -> str:
    """Return current Pacific Time timestamp in format YYYY-MM-DD HH:MM:SS."""
    utc_now = datetime.now(tz=ZoneInfo("UTC"))
    pacific_time = utc_now.astimezone(ZoneInfo("America/Los_Angeles"))
    return pacific_time.strftime("%Y-%m-%d %H:%M:%S")


def get_file_timestamp() -> str:
    """Return timestamp suitable for filenames: YYYY_MM_DD_HH_MM."""
    utc_now = datetime.now(tz=ZoneInfo("UTC"))
    pacific_time = utc_now.astimezone(ZoneInfo("America/Los_Angeles"))
    return pacific_time.strftime("%Y_%m_%d_%H_%M")


def parse_speed_to_bps(speed_str: str) -> int:
    """Convert speed string (e.g. '200Gbps', '1.6Tbps') to bps integer."""
    if not speed_str or speed_str == "Unknown":
        return 0
    clean = speed_str.strip()
    match = re.search(r"([\d\.]+)\s*([A-Za-z]+)", clean)
    if not match:
        return 0
    val = float(match.group(1))
    unit = match.group(2).lower()
    if "t" in unit:
        return int(val * 10**12)
    elif "g" in unit:
        return int(val * 10**9)
    elif "m" in unit:
        return int(val * 10**6)
    elif "k" in unit:
        return int(val * 10**3)
    return int(val)


def format_bps_human(bps: float) -> str:
    """Convert bps to human-readable string (bps, Kbps, Mbps, Gbps, Tbps)."""
    if bps >= 10**12:
        return f"{bps / 10**12:.2f} Tbps"
    elif bps >= 10**9:
        return f"{bps / 10**9:.2f} Gbps"
    elif bps >= 10**6:
        return f"{bps / 10**6:.2f} Mbps"
    elif bps >= 10**3:
        return f"{bps / 10**3:.2f} Kbps"
    return f"{bps:.0f} bps"


async def async_gnetch_command(cmd: str, target: str) -> list[str]:
    """Execute command on device asynchronously via Stubby gnetch-frontend."""
    stubby_cmd = f"stubby --proto2 call blade:gnetch-frontend Gnetch2.Command 'command: \"{cmd}\" target: \"{target}\"'"
    try:
        process = await asyncio.create_subprocess_shell(
            stubby_cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        stdout, _ = await process.communicate()
        raw_out = stdout.decode()

        match = re.search(r'data:\s*"(.*)"', raw_out, re.DOTALL)
        if match:
            content = match.group(1)
            content = content.replace('\\"', '"')
            lines = content.split("\\n")
            return [line.strip() for line in lines if line.strip()]
        
        if raw_out.strip() and not raw_out.startswith("timing_metrics"):
            lines = [l.strip() for l in raw_out.splitlines() if l.strip()]
            return lines

    except Exception:
        pass

    # Fallback to gnetch.sh
    if os.path.isfile(GNETCH_SH_PATH) and os.access(GNETCH_SH_PATH, os.X_OK):
        try:
            proc = await asyncio.create_subprocess_exec(
                GNETCH_SH_PATH, cmd, target,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )
            stdout, _ = await proc.communicate()
            output = stdout.decode().strip()
            return [l.strip() for l in output.splitlines() if l.strip()]
        except Exception:
            pass

    return []


async def rate_limited_gnetch_command(cmd: str, target: str) -> list[str]:
    """Execute command with per-device rate limiting."""
    if target not in rate_limiters:
        rate_limiters[target] = asyncio.Semaphore(1)

    async with rate_limiters[target]:
        output = await async_gnetch_command(cmd, target)
        await asyncio.sleep(RATE_LIMIT)
        return output


def parse_isis_adjacencies(lines: list[str]) -> list[dict]:
    """
    Parse 'show isis adjacency' output.
    Returns list of {'interface': 'aeX', 'subinterface': 'aeX.0', 'system': 'hostname', 'state': 'Up'}
    """
    adjacencies = []
    for line in lines:
        parts = line.strip().split()
        if len(parts) >= 4 and parts[0].startswith("ae"):
            sub_intf = parts[0]
            base_intf = sub_intf.split(".")[0]
            system = parts[1]
            # State is either parts[3] (when level is present) or parts[2]
            state = parts[3] if parts[2] in ["1", "2"] else parts[2]
            adjacencies.append({
                "interface": base_intf,
                "subinterface": sub_intf,
                "system": system,
                "state": state,
                "raw_line": line,
            })
    return adjacencies


def parse_interface_description_line(line: str) -> dict:
    """Parse a single output line from 'show interfaces descriptions | match ...'."""
    parts = line.split()
    if len(parts) < 3:
        return {
            "raw": line,
            "interface": parts[0] if parts else "",
            "admin": "",
            "link": "",
            "description": " ".join(parts[1:]) if len(parts) > 1 else "",
            "is_bundle": False,
            "is_subinterface": False,
            "bundle_parent": None,
            "remote_device": None,
        }

    intf_name = parts[0]
    admin = parts[1]
    link = parts[2]
    description = " ".join(parts[3:]) if len(parts) > 3 else ""

    is_bundle = intf_name.lower().startswith("ae")
    is_subinterface = "." in intf_name

    remote_match = re.search(r"\[R=([^\]]+)\]", description, re.IGNORECASE)
    remote_device = remote_match.group(1).strip() if remote_match else None

    agg_match = re.search(r"\[A=([^\]]+)\]", description, re.IGNORECASE)
    bundle_parent = agg_match.group(1).strip() if agg_match else None

    return {
        "raw": line,
        "interface": intf_name,
        "admin": admin,
        "link": link,
        "description": description,
        "is_bundle": is_bundle,
        "is_subinterface": is_subinterface,
        "bundle_parent": bundle_parent,
        "remote_device": remote_device,
    }


def parse_ae_interface_details(lines: list[str]) -> dict:
    """Parse 'show interface <ae_x>' output for capacity and live rates."""
    speed_str = "Unknown"
    speed_bps = 0
    input_bps = 0
    input_pps = 0
    output_bps = 0
    output_pps = 0
    admin_status = "Unknown"
    link_status = "Unknown"
    description = ""

    for line in lines:
        if line.startswith("Physical interface:"):
            if "Enabled" in line:
                admin_status = "Enabled"
            elif "Disabled" in line:
                admin_status = "Disabled"
            if "Physical link is Up" in line:
                link_status = "Up"
            elif "Physical link is Down" in line:
                link_status = "Down"

        elif "Description:" in line and not description:
            description = line.split("Description:", 1)[1].strip()

        elif "Speed:" in line:
            m_speed = re.search(r"Speed:\s*([0-9]+(?:Gbps|Mbps|kbps|bps))", line)
            if m_speed:
                speed_str = m_speed.group(1)
                speed_bps = parse_speed_to_bps(speed_str)

        elif "Input rate" in line:
            m_in = re.search(r"Input rate\s*:\s*([0-9]+)\s*bps\s*(?:\(([0-9]+)\s*pps\))?", line)
            if m_in:
                input_bps = int(m_in.group(1))
                if m_in.group(2):
                    input_pps = int(m_in.group(2))

        elif "Output rate" in line:
            m_out = re.search(r"Output rate\s*:\s*([0-9]+)\s*bps\s*(?:\(([0-9]+)\s*pps\))?", line)
            if m_out:
                output_bps = int(m_out.group(1))
                if m_out.group(2):
                    output_pps = int(m_out.group(2))

    return {
        "speed_str": speed_str,
        "speed_bps": speed_bps,
        "input_bps": input_bps,
        "input_pps": input_pps,
        "output_bps": output_bps,
        "output_pps": output_pps,
        "admin_status": admin_status,
        "link_status": link_status,
        "description": description,
        "raw_lines": lines,
    }


async def run_single_audit_target(
    device: str,
    target_neighbor: str,
    cmd_index: int,
    total_cmds: int
) -> dict:
    """
    Audit an adjacency pair:
      1. Source of Truth: 'show isis adjacency' on source router.
      2. Cross-reference: 'show interfaces descriptions | match <target_neighbor>'.
      3. Capacity & Rates: 'show interface <ae_x>' for each candidate AE.
    """
    # 1. Fetch ISIS Adjacencies (Source of Truth)
    isis_raw = await rate_limited_gnetch_command("show isis adjacency", device)
    all_isis_adjs = parse_isis_adjacencies(isis_raw)
    
    # Match ISIS adjacencies against target neighbor
    matched_isis_adjs = [
        adj for adj in all_isis_adjs
        if target_neighbor.lower() in adj["system"].lower()
    ]
    isis_ae_set = {adj["interface"] for adj in matched_isis_adjs if adj["state"].lower() == "up"}

    # 2. Fetch Interface Descriptions
    desc_cmd = f"show interfaces descriptions | match {target_neighbor}"
    desc_raw = await rate_limited_gnetch_command(desc_cmd, device)
    
    parsed_descriptions = []
    desc_ae_set = set()
    for line in desc_raw:
        parsed = parse_interface_description_line(line)
        parsed_descriptions.append(parsed)
        if parsed["is_bundle"] and not parsed["is_subinterface"]:
            desc_ae_set.add(parsed["interface"])

    # Combine AE candidates (ISIS adjacencies + description matches)
    all_ae_candidates = sorted(list(isis_ae_set.union(desc_ae_set)))

    # 3. Query Capacity and Traffic Rates for AE interfaces
    ae_details = {}
    for ae in all_ae_candidates:
        ae_raw = await rate_limited_gnetch_command(f"show interface {ae}", device)
        ae_parsed = parse_ae_interface_details(ae_raw)
        ae_parsed["interface"] = ae
        ae_parsed["in_isis"] = (ae in isis_ae_set)
        ae_parsed["in_desc"] = (ae in desc_ae_set)
        ae_details[ae] = ae_parsed

    # 4. Determine Active vs. Ignored (Drained / Idle / No-ISIS) AE Interfaces
    num_candidates = len(all_ae_candidates)
    active_aes = []
    ignored_aes = []

    for ae, details in ae_details.items():
        in_bps = details["input_bps"]
        out_bps = details["output_bps"]
        desc = details["description"].upper()

        is_traffic_idle = (in_bps < IDLE_TRAFFIC_THRESHOLD_BPS and out_bps < IDLE_TRAFFIC_THRESHOLD_BPS)
        is_drain_marked = ("DRAIN" in desc)
        not_in_isis = not details["in_isis"]

        if not_in_isis:
            details["status"] = "NO ISIS ADJACENCY (Ignored)"
            details["is_active"] = False
            ignored_aes.append(ae)
        elif num_candidates > 1 and (is_traffic_idle or is_drain_marked):
            details["status"] = "DRAINED/IDLE (Ignored)"
            details["is_active"] = False
            ignored_aes.append(ae)
        else:
            details["status"] = "ACTIVE"
            details["is_active"] = True
            active_aes.append(ae)

    # Compute Total Active Capacity and Rates
    total_active_capacity_bps = sum(ae_details[ae]["speed_bps"] for ae in active_aes)
    total_active_input_bps = sum(ae_details[ae]["input_bps"] for ae in active_aes)
    total_active_output_bps = sum(ae_details[ae]["output_bps"] for ae in active_aes)

    in_util_pct = (total_active_input_bps / total_active_capacity_bps * 100) if total_active_capacity_bps > 0 else 0.0
    out_util_pct = (total_active_output_bps / total_active_capacity_bps * 100) if total_active_capacity_bps > 0 else 0.0

    return {
        "index": cmd_index,
        "device": device,
        "target_neighbor": target_neighbor,
        "matched_isis_adjs": matched_isis_adjs,
        "desc_raw": desc_raw,
        "parsed_descriptions": parsed_descriptions,
        "ae_candidates": all_ae_candidates,
        "ae_details": ae_details,
        "active_aes": active_aes,
        "ignored_aes": ignored_aes,
        "total_active_capacity_bps": total_active_capacity_bps,
        "total_active_capacity_str": format_bps_human(total_active_capacity_bps) if total_active_capacity_bps > 0 else "0 bps",
        "total_active_input_bps": total_active_input_bps,
        "total_active_output_bps": total_active_output_bps,
        "in_util_pct": in_util_pct,
        "out_util_pct": out_util_pct,
    }


def export_report_to_excel(results: list[dict], output_file: str, timestamp_str: str):
    """
    Generate an Excel workbook with formatted sheets:
      Sheet 1: 'CR Capacity Summary'
      Sheet 2: 'AE Interfaces Detail'
      Sheet 3: 'Discovered Physical Links'
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Styling definitions
    font_title = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="595959")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_sub_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    font_data = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)
    fill_alt = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    fill_warn = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_alert = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # -------------------------------------------------------------
    # Sheet 1: CR Capacity Summary
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "CR Capacity Summary"

    ws1.merge_cells("A1:L1")
    ws1["A1"] = f"Core Router (CR) Capacity & Live Utilization Audit Report"
    ws1["A1"].font = font_title

    ws1.merge_cells("A2:L2")
    ws1["A2"] = f"Generated at: {timestamp_str} (Pacific Time) | Source of Truth: ISIS Adjacency & Interface Query"
    ws1["A2"].font = font_subtitle

    headers_ws1 = [
        "Device", "Neighbor Target", "Active AE(s)", "Ignored AE(s)",
        "Active Capacity", "Capacity (Gbps)", "Live Input Rate", "Input (Gbps)",
        "In Util (%)", "Live Output Rate", "Output (Gbps)", "Out Util (%)"
    ]

    ws1.row_dimensions[4].height = 24
    for col_idx, h in enumerate(headers_ws1, 1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    for row_idx, res in enumerate(results, 5):
        ws1.row_dimensions[row_idx].height = 19
        active_ae_str = ",".join(res["active_aes"]) if res["active_aes"] else "-"
        ignored_ae_str = ",".join(res["ignored_aes"]) if res["ignored_aes"] else "-"
        cap_gbps = res["total_active_capacity_bps"] / 10**9
        in_gbps = res["total_active_input_bps"] / 10**9
        out_gbps = res["total_active_output_bps"] / 10**9

        row_vals = [
            res["device"],
            res["target_neighbor"],
            active_ae_str,
            ignored_ae_str,
            res["total_active_capacity_str"],
            round(cap_gbps, 2),
            format_bps_human(res["total_active_input_bps"]),
            round(in_gbps, 2),
            round(res["in_util_pct"], 1),
            format_bps_human(res["total_active_output_bps"]),
            round(out_gbps, 2),
            round(res["out_util_pct"], 1),
        ]

        is_alt = (row_idx % 2 == 0)

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            if is_alt:
                cell.fill = fill_alt

            # Alignments & Formatting
            if col_idx in [1, 2]:
                cell.alignment = align_left
            elif col_idx in [3, 4]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right

            # Highlight high utilization
            if col_idx in [9, 12]:  # Util columns
                if isinstance(val, (int, float)):
                    if val >= 50.0:
                        cell.fill = fill_alert
                        cell.font = font_bold
                    elif val >= 30.0:
                        cell.fill = fill_warn

    # Auto-adjust column widths for Sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # -------------------------------------------------------------
    # Sheet 2: AE Interfaces Detail
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="AE Interfaces Detail")

    headers_ws2 = [
        "Device", "Neighbor Target", "AE Interface", "In ISIS Adjacency?",
        "In Descriptions?", "Capacity (Speed)", "Capacity (bps)",
        "Input Rate (Human)", "Input Rate (bps)", "Input (pps)",
        "Output Rate (Human)", "Output Rate (bps)", "Output (pps)",
        "Status", "Description"
    ]

    ws2.row_dimensions[1].height = 24
    for col_idx, h in enumerate(headers_ws2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_sub_header
        cell.alignment = align_center

    cur_r = 2
    for res in results:
        for ae, det in res["ae_details"].items():
            ws2.row_dimensions[cur_r].height = 18
            row_vals = [
                res["device"],
                res["target_neighbor"],
                ae,
                "Yes" if det["in_isis"] else "No",
                "Yes" if det["in_desc"] else "No",
                det["speed_str"],
                det["speed_bps"],
                format_bps_human(det["input_bps"]),
                det["input_bps"],
                det["input_pps"],
                format_bps_human(det["output_bps"]),
                det["output_bps"],
                det["output_pps"],
                det["status"],
                det["description"],
            ]
            for col_idx, val in enumerate(row_vals, 1):
                cell = ws2.cell(row=cur_r, column=col_idx, value=val)
                cell.font = font_data
                cell.border = thin_border
                if cur_r % 2 == 1:
                    cell.fill = fill_alt

                if col_idx in [1, 2, 14, 15]:
                    cell.alignment = align_left
                elif col_idx in [3, 4, 5]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_right

                if col_idx == 14:
                    if "ACTIVE" in str(val):
                        cell.font = font_bold
                    else:
                        cell.font = font_subtitle
            cur_r += 1

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(min(max_len + 3, 45), 12)

    # -------------------------------------------------------------
    # Sheet 3: Discovered Physical Links
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Discovered Physical Links")

    headers_ws3 = [
        "Device", "Neighbor Target", "Interface", "Admin State", "Link State",
        "Associated Bundle [A=]", "Remote Device [R=]", "Description / Circuit"
    ]

    ws3.row_dimensions[1].height = 24
    for col_idx, h in enumerate(headers_ws3, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    cur_r3 = 2
    for res in results:
        for item in res["parsed_descriptions"]:
            ws3.row_dimensions[cur_r3].height = 18
            row_vals = [
                res["device"],
                res["target_neighbor"],
                item["interface"],
                item["admin"],
                item["link"],
                item["bundle_parent"] or "-",
                item["remote_device"] or "-",
                item["description"],
            ]
            for col_idx, val in enumerate(row_vals, 1):
                cell = ws3.cell(row=cur_r3, column=col_idx, value=val)
                cell.font = font_data
                cell.border = thin_border
                if cur_r3 % 2 == 1:
                    cell.fill = fill_alt

                if col_idx in [1, 2, 8]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center
            cur_r3 += 1

    for col in ws3.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 12)

    wb.save(output_file)
    print(f"\n[Excel Export] Successfully saved full report to: {output_file}")


async def audit_cr_capacity(
    audit_targets: list[tuple[str, str]],
    deduplicate: bool = False,
    json_output_file: str = None,
    excel_output_file: str = None
) -> list[dict]:
    """Execute complete CR capacity audit across targets."""
    if deduplicate:
        seen = set()
        unique_targets = []
        for dev, tgt in audit_targets:
            key = (dev, tgt)
            if key not in seen:
                seen.add(key)
                unique_targets.append((dev, tgt))
        execution_list = unique_targets
    else:
        execution_list = audit_targets

    total_cmds = len(execution_list)
    ts_now = get_pacific_timestamp()
    print(f"\n{'=' * 95}")
    print(f"CR CAPACITY AUDIT: Authoritative ISIS Adjacency & Capacity Verification")
    print(f"Timestamp: {ts_now}")
    print(f"Total targets to audit: {total_cmds}")
    print(f"{'=' * 95}\n")

    # Run tasks concurrently
    tasks = [
        run_single_audit_target(dev, tgt, idx + 1, total_cmds)
        for idx, (dev, tgt) in enumerate(execution_list)
    ]
    results = await asyncio.gather(*tasks)
    results.sort(key=lambda x: x["index"])

    for res in results:
        idx = res["index"]
        device = res["device"]
        target = res["target_neighbor"]
        isis_adjs = res["matched_isis_adjs"]
        desc_raw = res["desc_raw"]
        parsed_desc = res["parsed_descriptions"]
        ae_details = res["ae_details"]
        active_aes = res["active_aes"]
        ignored_aes = res["ignored_aes"]

        print(f"\n[{idx}/{total_cmds}] Target: {device} ---> {target}")
        print(f"{'-' * 95}")

        # 1. ISIS Adjacency (Source of Truth)
        if isis_adjs:
            isis_summary = [f"{a['interface']} ({a['system']}, State: {a['state']})" for a in isis_adjs]
            print(f"  [1. ISIS Adjacency (Source of Truth)] Found {len(isis_adjs)} active adjacency:")
            for isis_s in isis_summary:
                print(f"    * {isis_s}")
        else:
            print("  [1. ISIS Adjacency (Source of Truth)] (No active ISIS adjacency found)")

        # 2. Interface Descriptions
        if desc_raw:
            print(f"  [2. Interface Descriptions] {len(desc_raw)} matching description line(s):")
            for line in desc_raw:
                print(f"    {line}")
            member_names = [i["interface"] for i in parsed_desc if not i["is_bundle"]]
            if member_names:
                print(f"    -> Member links: {member_names}")
        else:
            print("  [2. Interface Descriptions] (No matching interface descriptions)")

        # 3. AE Capacity & Traffic Rates
        if ae_details:
            print(f"  [3. AE Capacity & Traffic Rates]")
            for ae, det in ae_details.items():
                sp = det["speed_str"]
                in_str = format_bps_human(det["input_bps"])
                out_str = format_bps_human(det["output_bps"])
                in_pps = det["input_pps"]
                out_pps = det["output_pps"]
                status = det["status"]
                print(
                    f"    * {ae:<6} | Capacity: {sp:<10} | In: {in_str:<12} ({in_pps} pps) | "
                    f"Out: {out_str:<12} ({out_pps} pps) | Status: {status}"
                )

            if ignored_aes:
                print(f"  -> Ignored non-active/drained AE interface(s): {ignored_aes}")

            cap_str = res["total_active_capacity_str"]
            in_rate_str = format_bps_human(res["total_active_input_bps"])
            out_rate_str = format_bps_human(res["total_active_output_bps"])
            print(
                f"  => Active Capacity: {cap_str} | Active In: {in_rate_str} ({res['in_util_pct']:.1f}%) | "
                f"Active Out: {out_rate_str} ({res['out_util_pct']:.1f}%)"
            )
        else:
            print("  [3. AE Capacity] No candidate AE interfaces found.")

    # Consolidated Capacity Summary Table
    print(f"\n{'=' * 95}")
    print("CONSOLIDATED CORE ROUTER (CR) CAPACITY & UTILIZATION SUMMARY")
    print(f"{'=' * 95}")
    header = (
        f"{'Device':<13} {'Neighbor Target':<16} {'Active AE':<10} {'Ignored AE':<12} "
        f"{'Capacity':<11} {'Input Rate':<13} {'In Util':<9} {'Output Rate':<13} {'Out Util'}"
    )
    print(header)
    print(f"{'-' * 95}")

    for res in results:
        device = res["device"]
        target = res["target_neighbor"]
        active_ae_str = ",".join(res["active_aes"]) if res["active_aes"] else "-"
        ignored_ae_str = ",".join(res["ignored_aes"]) if res["ignored_aes"] else "-"
        cap_str = res["total_active_capacity_str"]
        in_rate_str = format_bps_human(res["total_active_input_bps"])
        out_rate_str = format_bps_human(res["total_active_output_bps"])
        in_util_str = f"{res['in_util_pct']:.1f}%" if res["total_active_capacity_bps"] > 0 else "-"
        out_util_str = f"{res['out_util_pct']:.1f}%" if res["total_active_capacity_bps"] > 0 else "-"

        print(
            f"{device:<13} {target:<16} {active_ae_str:<10} {ignored_ae_str:<12} "
            f"{cap_str:<11} {in_rate_str:<13} {in_util_str:<9} {out_rate_str:<13} {out_util_str}"
        )

    print(f"{'-' * 95}")
    print(f"Total audit targets processed: {total_cmds}")
    print(f"{'=' * 95}\n")

    if json_output_file:
        export_data = {
            "audit_timestamp": ts_now,
            "total_commands": total_cmds,
            "results": results,
        }
        with open(json_output_file, "w", encoding="utf-8") as f:
            json.dump(export_data, f, indent=2, default=str)
        print(f"[JSON Export] Saved audit capacity result to: {json_output_file}\n")

    if excel_output_file:
        export_report_to_excel(results, excel_output_file, ts_now)

    return results


def parse_args():
    parser = argparse.ArgumentParser(
        description="Audit CR capacity using source router ISIS adjacency as source of truth and export reports."
    )
    parser.add_argument(
        "--dedup",
        action="store_true",
        help="Deduplicate identical (device, target) pairs before running",
    )
    parser.add_argument(
        "-e", "--excel",
        nargs="?",
        const="",
        default=None,
        help="Export full report to Excel workbook (.xlsx). Defaults to 'cr_capacity_audit_<timestamp>.xlsx' if no filename is specified."
    )
    parser.add_argument(
        "-o", "--output-json",
        default=None,
        help="Optional JSON filename to export capacity audit data",
    )
    parser.add_argument(
        "-d", "--device",
        default=None,
        help="Filter to only run on a specific device (e.g. cr01.atl101)",
    )
    parser.add_argument(
        "-t", "--target",
        default=None,
        help="Filter to only audit a specific neighbor target (e.g. cr01.aus122)",
    )
    return parser.parse_args()


async def main():
    args = parse_args()

    targets = DEFAULT_AUDIT_COMMANDS

    if args.device:
        targets = [item for item in targets if args.device.lower() in item[0].lower()]

    if args.target:
        targets = [item for item in targets if args.target.lower() in item[1].lower()]

    if not targets:
        print("No audit targets matched the specified filters.")
        return

    excel_filename = None
    if args.excel is not None:
        if args.excel == "":
            excel_filename = f"cr_capacity_audit_{get_file_timestamp()}.xlsx"
        else:
            excel_filename = args.excel
            if not excel_filename.endswith(".xlsx"):
                excel_filename += ".xlsx"

    await audit_cr_capacity(
        targets,
        deduplicate=args.dedup,
        json_output_file=args.output_json,
        excel_output_file=excel_filename,
    )


if __name__ == "__main__":
    asyncio.run(main())
