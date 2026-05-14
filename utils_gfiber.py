#import telnetlib
import ipaddress
import socket
import sys
import argparse
import time
import logging
import traceback
import pprint
#import paramiko
import time
from time import sleep
import re
import os
import json
from datetime import datetime
import xlsxwriter
from excel import *
#from ixia_ngfp_lib import *
import settings
from console_util  import  *
import pexpect
from threading import Thread
import subprocess
import shutil
#import spur
import xmlformatter
import xml.etree.ElementTree as ET
import yaml
import shutil
import asyncio
import json
from openpyxl.styles import Font


if sys.platform.startswith('linux'):
    import pexpect
elif sys.platform.startswith('win'):
    import wexpect
elif sys.platform.startswith('darwin'):
    # Code for macOS platform
    pass
else:
    raise OSError('Unsupported operating system')

# Use pexpect or wexpect as needed


class CustomArgumentParser(argparse.ArgumentParser):
    def print_help(self):
        super().print_help()
        print("Note #1: Make sure you fill up the network_saos_10.ymal file. ")
        print("Note #2: If you want to use MCP exported file, the file has to be csv, not xlsl")
        print("Note #3: If you don't provide any arugment in CLI, by default you are using SSH to access devices")


DEBUG = False
INFO = True 

RATE_LIMIT = 1  # seconds between commands for the same device
rate_limiters = {}

async def async_gnetch_command(cmd, target):
    """
    Asynchronously execute a gnetch command using asyncio subprocess.
    """
    gnetch_cmd = f"stubby --proto2 call blade:gnetch-frontend Gnetch2.Command 'command: \"{cmd}\" target: \"{target}\"'"
    process = await asyncio.create_subprocess_shell(
        gnetch_cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE
    )
    stdout, _ = await process.communicate()
    output = stdout.decode().lstrip("data: ").strip('"').split("\\n")
    return output[:-1]

async def rate_limited_gnetch_command(cmd, target):
    """
    Execute gnetch_command with rate limiting for the target device.
    """
    if target not in rate_limiters:
        rate_limiters[target] = asyncio.Semaphore(1)

    async with rate_limiters[target]:
        output = await async_gnetch_command(cmd, target)
        await asyncio.sleep(RATE_LIMIT)  # Enforce the rate limit
        return output

def load_json_file(json_file_path):
	try:
	    # Open the file and load its content into a dictionary.
	    with open(json_file_path, "r") as file:
	        audit_ipfix_dict = json.load(file)
	    print("JSON file loaded successfully.")
	except FileNotFoundError:
	    print(f"Error: The file '{json_file_path}' was not found.")
	    audit_ipfix_dict = {}
	except json.JSONDecodeError as error:
	    print(f"Error: Could not decode JSON. {error}")
	    audit_ipfix_dict = {}
	return audit_ipfix_dict

def human_readable_format(value, unit="bps"):
    """
    Convert a large float value into a human-readable format with K, M, G, T suffixes.
    
    Parameters:
    - value (float): The numerical value to be converted.
    - unit (str): The base unit (default is "bps" for bits per second).

    Returns:
    - str: Human-readable format (e.g., "15.0Gbps").
    """
    # Define unit multipliers
    suffixes = ["", "K", "M", "G", "T", "P"]  # Supports up to Petabits
    factor = 1000.0  # Conversion factor (1K = 1000)

    for suffix in suffixes:
        if abs(value) < factor:
            return f"{value:.2f}{suffix}{unit}"
        value /= factor  # Reduce value by a factor of 1000

    return f"{value:.2f}P{unit}"  # If value is extremely large, default to Petabits


def convert_to_bps(speed_str):
    """
    Convert a string with a speed value and unit (e.g., '18.6769Gbps', '100.0bps') into a float value in bits per second (bps).
    
    Supports the following units:
    - bps          -> Bits per second (1 bps = 1 bps)
    - Kbps / kbps  -> Kilobits per second (1 Kbps = 1,000 bps)
    - Mbps         -> Megabits per second (1 Mbps = 1,000,000 bps)
    - Gbps         -> Gigabits per second (1 Gbps = 1,000,000,000 bps)
    - Tbps         -> Terabits per second (1 Tbps = 1,000,000,000,000 bps)
    """
    # Define a regex pattern to extract the numeric value and unit
    pattern = r"([\d\.]+)([KkMmGgTt]?bps)"

    match = re.match(pattern, speed_str.strip())
    if not match:
        raise ValueError(f"Invalid format: {speed_str}")

    # Extract the numeric value and unit
    value, unit = match.groups()
    value = float(value)  # Convert the numeric part to float

    # Define unit multipliers (bps as the base unit)
    unit_multipliers = {
        "bps": 1,           # No conversion needed
        "Kbps": 1_000,
        "kbps": 1_000,
        "Mbps": 1_000_000,
        "Gbps": 1_000_000_000,
        "Tbps": 1_000_000_000_000
    }

    # Convert to bps
    if unit in unit_multipliers:
        return value * unit_multipliers[unit]
    else:
        raise ValueError(f"Unknown unit: {unit}")


def dump_json_file(folder_name,file_name,dict_data):

	# Ensure the target folder exists; if not, create it
	if not os.path.exists(folder_name):
	    os.makedirs(folder_name)

	# Full path for the output JSON file
	file_path = os.path.join(folder_name, file_name)

	# Write the dictionary to the JSON file
	try:
	    with open(file_path, "w") as json_file:
	        json.dump(dict_data, json_file, indent=4)
	    print(f"Dictionary data successfully dumped to '{file_path}'")
	except Exception as e:
	    print(f"An error occurred: {e}")

def create_or_recreate_folder(folder_name):
    """
    Creates a folder if it doesn't exist. If it exists, it deletes all files and subdirectories inside it,
    leaving the folder intact.

    Args:
        folder_name: The name (or path) of the folder.
    """
    if os.path.exists(folder_name):
        # Delete all files and directories within the folder without deleting the folder itself.
        for filename in os.listdir(folder_name):
            file_path = os.path.join(folder_name, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)  # Remove file or symbolic link.
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)  # Remove directory and its contents.
            except Exception as e:
                print(f"Failed to delete {file_path}. Reason: {e}")
    else:
        os.makedirs(folder_name)

def create_or_recreate_folder_old(folder_name):
	"""
	Creates a folder if it doesn't exist. If it exists, it deletes it and recreates it.

	Args:
	folder_name: The name of the folder to create.
	"""
	if os.path.exists(folder_name):
		shutil.rmtree(folder_name)  # Delete the existing folder
	os.makedirs(folder_name)  # Create the folder

def remove_all_files_in_folder(folder_path):
    if not os.path.exists(folder_path):
        print(f"The folder '{folder_path}' does not exist.")
        return
    
    # Remove all files in the folder
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)  # Remove file or symlink
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)  # Remove directory and its contents
        except Exception as e:
            print(f"Failed to delete {file_path}. Reason: {e}")
    
    print(f"All files in the folder '{folder_path}' have been removed.")
 
def write_list_to_file(filename, text_list):
    print(f"Writing to {filename}....")
    with open(filename, 'w') as file:
        file.write('\n'.join(text_list))

def jinja_zip(*args):
	return zip(*args)

def list_files(directory):
    file_list = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            file_list.append(file_path)
    return file_list


def horizontal_worksheet_title(horizontal_title_list,ws):
	ws.column_dimensions['A'].width = 15
	ws.column_dimensions['B'].width = 7
	ws.column_dimensions['C'].width = 15
	ws.column_dimensions['D'].width = 15
	ws.column_dimensions['E'].width = 15
	ws.column_dimensions['F'].width = 8
	ws.column_dimensions['G'].width = 15
	ws.column_dimensions['H'].width = 20
	ws.column_dimensions['I'].width = 30
	ws.column_dimensions['J'].width = 10
	

	for column, title in enumerate(horizontal_title_list):
		ws.cell(row=1, column=column+1, value=title)

def update_horizontal_ws(ws,*args,**kwargs):
	column = kwargs["column"]
	last_row = kwargs["last_row"]
	dprint(f"in update_horizontal_ws: last_row = {last_row}")
	value =kwargs["value"]
	if type(value) == list:
		for i in range(len(value)):
			ws[f"{column}{last_row+i+1}"] = value[i]
	else:
		ws[f"{column}{last_row+1}"] = value
	

def check_list(l):
    for prev, next in zip(l[:-1],l[1:]):
        if set(prev) != set(next):
            return False
    return True
    
def dump_dict_yaml_file(file,d):
	with open(file, 'w') as f:
		yaml.dump(d, f)

def xml_format(file):
	formatter = xmlformatter.Formatter(indent="1", indent_char="\t", encoding_output="ISO-8859-1", preserve=["literal"])
	formatter.format_file(file)


def send_Message(stock_msg):
  #stock_msg = remove_bracket(stock_msg)
  cmd = """osascript send_imessage.applescript 4088967681 '{}' """.format(stock_msg)
  #print(new_cmd)
  os.system(cmd)
  return None

def write_file(file_name,content):
	with open(file_name, 'w') as f:
		f.write(content)

def xml_file_string(file):
	tree = ET.parse(file)
	root = tree.getroot()
	xmlstr = ET.tostring(root, encoding='utf8', method='xml')
	return xmlstr 

def write_file_xml(file_name,content):
	with open(file_name, 'w') as f:
		f.write(content)
	formatter = xmlformatter.Formatter(indent="1", indent_char="\t", encoding_output="ISO-8859-1", preserve=["literal"])
	formatter.format_file(file_name)
 
def ip_break_up(ip):
	matched = re.search(r'([0-9]+\.[0-9]+\.[0-9]+)\.([0-9]+)',ip)
	if matched:
	  net = matched.group(1)
	  host = matched.group(2)
	  return net,host
	return None, None


def list_add(proc_list_all,proc_list):
	for i in proc_list:
		proc_list_all.append(i)

def init_tracking_loop(loop_count):
	loop_count = 0

def tracking_loop(loop_count,mac_list):
	loop_count +=1
	if loop_count == len(mac_list):
		Tracking = "End"
	elif loop_count == len(mac_list) - 1:
		Tracking = "Penultimate"
	elif loop_count == 1:
		Tracking = "Start"
	else:
		Tracking = "Middle"
	return Tracking

def touch(fname):
    if os.path.exists(fname):
        os.utime(fname, None)
    else:
        open(fname, 'a').close()

def print_bytes(bytes):
	message = bytes.strip()
	lines = message.splitlines()
	for line in lines:
		print (line.decode('ascii'))

def scp_file(**kwargs):
	if "file" in kwargs:
		FILE = kwargs['file']
	else:
		FILE="MCLAG_Perf_448D_1.xlsx"
	if "server" in kwargs:
		HOST=kwargs['server']
	else:
		HOST="10.105.19.19"
	if "password" in kwargs:
		PASS=kwargs['password']
	else:
		PASS="Shenghuo2014+"
	if "user" in kwargs:
		USER=kwargs['user']
	else:
		USER="zhengh"
	REMOTE_FILE=""
	
	
	COMMAND="scp -oPubKeyAuthentication=no %s %s@%s:%s" % (FILE, USER, HOST, REMOTE_FILE)

	child = pexpect.spawn(COMMAND)
	child.expect('password:')
	child.sendline(PASS)
	child.expect(pexpect.EOF)
	print_bytes(child.before)

class Logger(object):
    def __init__(self, file):
        self.terminal = sys.stdout
        print(f"Log file = {file}")
        self.log = open(file, "w")

    def write(self, message):
        try:
            self.terminal.write(message)
            self.terminal.flush()
            self.log.write(message)
            self.log.flush()
        except UnicodeEncodeError as e:
            print("Error: UnicodeEncodeError")

    def flush(self):
        self.terminal.flush()

    def close(self):
        # self.terminal.close()
        self.log.close()

# def tprint(var):
#     print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))+" :: "+str(var))



def parse_sys_top(result):
	high_lines = []
	high = False
	cpu_dict={}
	whole = ''
	busy_dict = {}
	for line in result:
		if "U" in line and "S" in line and "I" in line:
			debug("parse_sys_top:found cpu utils headline")
			items = re.split(';\\s+|,\\s+|\n',line)
			debug(f"(parse_sys_top: headline parsing: {items}")
			user = items[0]
			user = int((re.match('([0-9]+)(U)',user)).group(1))
			system = items[1]
			system = int((re.match('([0-9]+)(S)',system)).group(1))
			idle = items[2]
			idle = int(re.match('([0-9]+)(I)',idle).group(1))
			busy_dict['user'] = user
			busy_dict['system'] = system
			busy_dict['idle'] = idle
			whole = line
			debug(f'parse_sys_top: headline = {whole}')
			continue
			# print(user,system,idle)
		if line == '':
			continue
		line = line.strip()
		items = re.split('\\s+|,\\s+|\n',line)
		debug(items)
		if len(items) != 5 and len(items) !=6 :
			continue
		try:
			debug("parse_sys_top: items length = {}".format(len(items)))
			if len(items) == 5:
				obj = float(items[3])
				debug("parse_sys_top: items[3] = {}".format(obj))
			if len(items) == 6:
				obj = float(items[4])
				debug("parse_sys_top: items[4] = {}".format(obj))
			cpu_dict[items[0]] = {}
			cpu_dict[items[0]]['cpu'] = obj
			cpu_dict[items[0]]['line'] = line
			cpu_dict[items[0]]['headline'] = whole
			cpu_dict[items[0]]['headline_dict'] = busy_dict

			if obj > 30.0:
				high_lines.append(line)
				debug("parse_sys_top: cpu is high")
				high = True
		except Exception as e:
			tprint("parse_sys_top: line not parsable")
	debug (high_lines)
	debug (cpu_dict)
	return (high,high_lines,cpu_dict)

def print_cmd_output(msg,dut_name,cmd):
	# global DEBUG
	# print(DEBUG)
	tprint("========== Commnd output at {}: {}".format(dut_name,cmd))
	if type(msg) == list:
		for m in msg:
			tprint("{}: {}".format(dut_name,m))
	else:
		tprint("{}: {}".format(dut_name,msg))

def print_cmd_output_from_list(msgs):
	# global DEBUG
	# print(DEBUG)
	if type(msgs) == list:
		for m in msgs:
			tprint(f"{m}")
	else:
		tprint(f"{msgs}")

def print_file(msg, file,**kwargs):
	if "dut_name" in kwargs:
		dut_name = kwargs['dut_name']
	else:
		dut_name = "DUT"
	with open(file,'a+') as f:
		if type(msg) == list:
			for m in msg:
				f.write(time_str("{}:{}\n".format(dut_name,m)))
		else:
			f.write(time_str("{}:{}\n".format(dut_name,msg)))

def print_dict(d, indent=0):
    for key, value in d.items():
        print('  ' * indent + str(key) + ':', end=' ')
        if isinstance(value, dict):
            print('')
            print_dict(value, indent + 1)
        else:
            print(value)

def print_dict_original(d):
	print(yaml.dump(d, default_flow_style=False, explicit_start=True,sort_keys=False))
	# pp = pprint.PrettyPrinter(depth=4)
	# pp.pprint(d)

def dprint(*args):
    if DEBUG:
        for msg in args:
            if type(msg) == list:
                for m in msg:
                    tprint("Debug: {}".format(m))
            else:
                tprint("Debug: {}".format(msg))

def delete_file(file):
	if os.path.exists(file):
		os.remove(file)
		print(f"Deleted file: {file}")

def delete_files_in_directory(directory):
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            os.remove(file_path)
            print(f"Deleted file: {file_path}")


def delete_folder_contents(folder_path):
    # Delete all contents of the folder
    try:
    	shutil.rmtree(folder_path)
    	print(f"All contents of {folder_path} deleted.")
    except Exception as e:
    	print(f"Folder {folder_path} does not exsit")

def input_value(input_msg,default=""):
	value =input(input_msg)
	if value == "":
		return default 
	else:
		return value 

def delete_files_in_folder(folder_path):
    # Iterate over all items in the folder
    for item in os.listdir(folder_path):
        item_path = os.path.join(folder_path, item)
        # If it's a file, delete it
        if os.path.isfile(item_path):
            os.remove(item_path)
            print(f"Deleted file: {item_path}")
        # If it's a directory, recursively call the function
        elif os.path.isdir(item_path):
            delete_files_in_folder(item_path)
            # Once all contents are deleted, remove the directory itself
            os.rmdir(item_path)
            print(f"Deleted directory: {item_path}")

def print_output_list(msg):
	if type(msg) == list:
		for m in msg:
			tprint(f"{m}")
	else:
		tprint(f"Debug: {msg}")

def debug(msg):
	# global DEBUG
	#print(f"DEBUG Mode = {settings.DEBUG}")
	if DEBUG:
		if type(msg) == list:
			for m in msg:
				tprint("Debug: {}".format(m))
		else:
			tprint("Debug: {}".format(msg))

def kwargs_args_example(*args, **kwargs):

	for num in args:
		z+= num 

	tkwargs = {}
	for k,v in kwargs.items():
		tkwargs[k] = v
		if k == "final":
			final = tkwargs[k]
			break
	z += final 
	return z

def time_str(mesg):
	return str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " :: " + mesg

def tprint(*args, **kwargs):
    tempa = ' '.join(str(a) for a in args)
    tempk = ' '.join([str(kwargs[k]) for k in kwargs])
    temp = tempa + ' ' + tempk # puts a space between the two for clean output
    print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " :: " + temp)

def Info(*args, **kwargs):
    if INFO != True:
	    return
    tempa = ' '.join(str(a) for a in args)
    tempk = ' '.join([str(kwargs[k]) for k in kwargs])
    temp = tempa + ' ' + tempk # puts a space between the two for clean output
    print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " :: " + "Info: " + temp)
    print('\n')

def ErrorNotify(*args, **kwargs):
    tempa = ' '.join(str(a) for a in args)
    tempk = ' '.join([str(kwargs[k]) for k in kwargs])
    temp = tempa + ' ' + tempk # puts a space between the two for clean output
    print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " :: " + "Error: " + temp)

def WarningNotify(*args, **kwargs):
    tempa = ' '.join(str(a) for a in args)
    tempk = ' '.join([str(kwargs[k]) for k in kwargs])
    temp = tempa + ' ' + tempk # puts a space between the two for clean output
    print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " :: " + "Warning: " + temp)


def threads_exit(stop_threads,threads_list):
	# stop_threads = True
	for t in threads_list:
		t.join()

def convert_cmd_ascii(cmd):
	return cmd.encode('ascii')

def convert_cmd_ascii_n(cmd):
	cmd = cmd + '\n'
	return cmd.encode('ascii')

def relogin_dut_all(dut_list):
	for dut in dut_list:
		relogin_if_needed(dut)

	config_admin_timeout(dut_list)

def relogin_after_reboot(dut):
	time.sleep(200)
	relogin_if_needed(dut)

def switch_exec_reboot(dut,**kwargs):
	if "device" in kwargs:
		dev_name = kwargs['device']
	else:
		dev_name = "DUT"

	tprint("-------- Rebooting device : {}".format(dev_name))
	switch_interactive_exec(dut,"exec reboot","Do you want to continue? (y/n)")
	# thread = Thread(target = relogin_after_reboot,args = (dut,))
	# thread.start()
	return

def switch_flap_port(tn, port):
	switch_shut_port(tn,port)
	sleep(3)
	switch_unshut_port(tn,port)

  
def send_ctrl_c_cmd(tn):
	tn.write(('\x03').encode('ascii'))

def collect_edit_question_cmd(tn,cmd,**kwargs):
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 3
	#relogin_if_needed(tn)
	cmd = convert_cmd_ascii(cmd)
	tn.write(cmd)
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	sleep(timeout)
	output = tn.read_very_eager()
	#output = tn.read_until(("# ").encode('ascii'))
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).rstrip(' ')
		out_str_list.append(o_str)
	# tprint(dir(output))
	# tprint(type(output))
	#tprint(out_list)
	# for i in out_str_list:
	# 	tprint(i)
	return out_str_list

def clean_show_output(out_str_list,cmd):
	i = 0
	i_list = []
	for o in out_str_list:
		if str(cmd) in str(o):
			i_list.append(i)
		i += 1
	index = i_list[-1]
	good_out_list = out_str_list[index:]
	debug(good_out_list)
	return good_out_list

def clean_show_output_recursive_general(out_str_list,cmd_list):
	for cmd in cmd_list:
		for o in out_str_list:
			if str(cmd) in str(o):
				return out_str_list
			else:
				out_str_list.pop(0)
				return clean_show_output_recursive(out_str_list,cmd)

def clean_show_output_recursive(out_str_list,cmd):
	for o in out_str_list:
		if str(cmd) in str(o):
			return out_str_list
		else:
			out_str_list.pop(0)
			return clean_show_output_recursive(out_str_list,cmd)

def read_console(tn):
	sleep(5)
	output = tn.read_very_eager()
	print(output)
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)
		 
	print(out_str_list)
	return out_str_list

def print_show_cmd_list_generic(tn,cmd_f_string,*args,**kwargs):
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 5
	#relogin_if_needed(tn)
	if "logger" in kwargs:
		mylogger = kwargs["logger"]
	else:
		mylogger = None
	handle_prompt_before_commands(tn)
	original_cmds = cmd_f_string
	cmd_list = split_fstring_lines_generic(cmd_f_string)	
	for cmd in cmd_list:
		cmd_bytes = convert_cmd_ascii_n(cmd)
		tn.write(cmd_bytes)
	sleep(timeout)
	output = tn.read_very_eager()
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)
	for out_str in out_str_list:
		tprint(f"{str(out_str)}\n")
		if mylogger != None:
			mylogger.write(f"{str(out_str)}\n")

def print_show_cmd(tn,cmd,*args,**kwargs):
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 5
	#relogin_if_needed(tn)
	if "logger" in kwargs:
		mylogger = kwargs["logger"]
	else:
		mylogger = None
	if "mode" in kwargs:
		mode = kwargs["mode"]
	else:
		mode = "slow"
	Info(f"print_show_cmd: mode = {mode}")
	if mode == "slow":
		handle_prompt_before_commands(tn)
	original_cmd = cmd
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(cmd_bytes)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	sleep(timeout)
	output = tn.read_very_eager()
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)
	for out_str in out_str_list:
		tprint(f"{str(out_str)}\n")
		if mylogger != None:
			mylogger.write(f"{str(out_str)}\n")

def collect_show_cmd_general(tn,*args,**kwargs):
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 5
	#relogin_if_needed(tn)
	origal_cmd_list = args
	handle_prompt_before_commands(tn)
	for cmd in args:
		original_cmd = cmd
		cmd_bytes = convert_cmd_ascii_n(cmd)
		tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
		tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
		tn.write(cmd_bytes)
		tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
		tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
		sleep(timeout)
	output = tn.read_very_eager()
	#print(output)
	#output = tn.read_until(("# ").encode('ascii'))
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)
	good_out_list = clean_show_output_recursive_general(out_str_list,origal_cmd_list)
	debug(good_out_list)
	print_output_list(good_out_list)
	return good_out_list

def ciena_split_long_line(line):
 	items = [i.strip() for i in line.split("|")]
 	return items 

def ciena_split_line(line):
	items = line.split("|")
	return(items[1].strip(),items[2].strip())

def simple_input_manual(msg = ""):
	while True:
		answer = input(f"{msg} Or you want to quit(Y/N/Q):")
		if answer.upper() == "Y":
			return "Y"
		elif answer.upper() == "N":
			return "N"
		elif answer.upper() == "Q":
			return "Q"



def input_yes(msg = ""):
	answer = input(f"{msg}(Y/N):")
	if answer.upper() == "Y":
		return True 
	else:
		return False 

def collect_show_cmd_ciena(tn,cmd,**kwargs):
	empty_list = []
	if tn == None:
		print(f"telnet handle is none, return")
		return empty_list
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 2
	if 'mode' in kwargs:
		mode = kwargs['mode']
	else:
		mode = "fast"
	if "ssh" in kwargs:
		ssh = kwargs['ssh']
		return ssh.cmd_proc(cmd)
	print(cmd)
	debug(f"At collect_show_cmd: mode = {mode}")
	#relogin_if_needed(tn)
	if mode == "slow":
		handle_prompt_before_commands(tn)
	original_cmd = cmd
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(cmd_bytes)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	sleep(timeout)
	try:
		#output = tn.read_until(("> ").encode('ascii'))
		output = tn.read_very_eager()
	except Exception as ConnectionResetError:
		return "ConnectionResetError"
	except Exception as e:
		output = tn.read_very_eager()
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'

	out_str_list = []
	empty_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		if o_str !="" or cmd not in o_str:
			out_str_list.append(o_str)
	try: 
		for msg in out_str_list:
			print(msg)
		return out_str_list
	except Exception as e:
		return empty_list


def collect_show_cmd(tn,cmd,**kwargs):
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 2
	if 'mode' in kwargs:
		mode = kwargs['mode']
	else:
		mode = "fast"
	if "ssh" in kwargs:
		ssh = kwargs['ssh']
		return ssh.cmd_proc(cmd)

	debug(f"At collect_show_cmd: mode = {mode}")
	#relogin_if_needed(tn)
	if mode == "slow":
		handle_prompt_before_commands(tn)
	original_cmd = cmd
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(cmd_bytes)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	sleep(timeout)
	output = tn.read_very_eager()
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)

	good_out_list = clean_show_output_recursive(out_str_list,original_cmd)
	debug(good_out_list)
	print_output_list(good_out_list)
	return good_out_list

def log_show_cmd_telnet(tn,cmd,**kwargs):
	empty_list = []
	if tn == None:
		print(f"telnet handle is none, return")
		return empty_list
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 2
	if 'mode' in kwargs:
		mode = kwargs['mode']
	else:
		mode = "fast"
	if "ssh" in kwargs:
		ssh = kwargs['ssh']
		return ssh.cmd_proc(cmd)
	
	debug(f"At collect_show_cmd: mode = {mode}")
	#relogin_if_needed(tn)
	if mode == "slow":
		handle_prompt_before_commands(tn)
	original_cmd = cmd
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(cmd_bytes)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	sleep(timeout)
	try:
		#output = tn.read_until(("# ").encode('ascii'))
		output = tn.read_very_eager()
	except Exception as ConnectionResetError:
		return "ConnectionResetError"
	except Exception as e:
		output = tn.read_very_eager()
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)

	good_out_list = clean_show_output_recursive(out_str_list,original_cmd)
	if good_out_list == None:
		return empty_list
	try: 
		for msg in good_out_list:
			print(msg)
		return good_out_list
	except Exception as e:
		return empty_list

def log_show_cmd_telnet_6500(tn,cmd,**kwargs):
	empty_list = []
	if tn == None:
		print(f"telnet handle is none, return")
		return empty_list
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 2
	if 'mode' in kwargs:
		mode = kwargs['mode']
	else:
		mode = "fast"
	
	debug(f"At collect_show_cmd: mode = {mode}")
	#relogin_if_needed(tn)
	if mode == "slow":
		handle_prompt_before_commands(tn)
	original_cmd = cmd
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(cmd_bytes)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	sleep(timeout)
	try:
		#output = tn.read_until(("# ").encode('ascii'))
		#tn.set_debuglevel(10000000)
		output = tn.read_very_eager()
	except Exception as ConnectionResetError:
		return "ConnectionResetError"
	except Exception as e:
		output = tn.read_very_eager()
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)

	good_out_list = out_str_list
	if good_out_list == None:
		return empty_list
	try: 
		for msg in good_out_list:
			print(msg)
		return good_out_list
	except Exception as e:
		return empty_list

def exec_cmd_telnet(tn,cmd,**kwargs):
	empty_list = []
	if tn == None:
		print(f"telnet handle is none, return")
		return empty_list
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 2
	if 'mode' in kwargs:
		mode = kwargs['mode']
	else:
		mode = "fast"
	if "ssh" in kwargs:
		ssh = kwargs['ssh']
		return ssh.cmd_proc(cmd)
	#print(cmd)
	debug(f"At collect_show_cmd: mode = {mode}")
	#relogin_if_needed(tn)
	if mode == "slow":
		handle_prompt_before_commands(tn)
	original_cmd = cmd
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(cmd_bytes)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	sleep(timeout)
	try:
		#output = tn.read_until(("> ").encode('ascii'))
		output = tn.read_very_eager()
	except Exception as ConnectionResetError:
		return "ConnectionResetError"
	except Exception as e:
		output = tn.read_very_eager()
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)

	good_out_list = clean_show_output_recursive(out_str_list,original_cmd)
	if good_out_list == None:
		return empty_list
	try: 
		for msg in good_out_list:
			print(msg)
		return good_out_list
	except Exception as e:
		return empty_list

def exec_cmd_telnet_onu(tn,cmd,**kwargs):
	empty_list = []
	if tn == None:
		print(f"telnet handle is none, return")
		return empty_list
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 2
	if 'mode' in kwargs:
		mode = kwargs['mode']
	else:
		mode = "fast"
	if "ssh" in kwargs:
		ssh = kwargs['ssh']
		return ssh.cmd_proc(cmd)
	print(cmd)
	debug(f"At collect_show_cmd: mode = {mode}")
	#relogin_if_needed(tn)
	if mode == "slow":
		handle_prompt_before_commands(tn)
	original_cmd = cmd
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(cmd_bytes)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	sleep(timeout)
	try:
		#output = tn.read_until(("> ").encode('ascii'))
		output = tn.read_very_eager()
	except Exception as ConnectionResetError:
		return "ConnectionResetError"
	except Exception as e:
		output = tn.read_very_eager()
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'

	out_str_list = []
	empty_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		if o_str !="" or cmd not in o_str:
			out_str_list.append(o_str)
	try: 
		for msg in out_str_list:
			print(msg)
		return out_str_list
	except Exception as e:
		return empty_list


def show_execute_cmd(tn,cmd,**kwargs):
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 8
	#relogin_if_needed(tn)
	original_cmd = cmd
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(cmd_bytes)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	sleep(timeout)
	output = tn.read_very_eager()
	debug(output)
	#output = tn.read_until(("# ").encode('ascii'))
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)
	 
	if cmd == "get router info6 bgp summary":
		print (f"return from utiliy.py: collect_show_cmd(): {out_str_list}")

	good_out_list = clean_show_output_recursive(out_str_list,original_cmd)
	print_cmd_output_from_list(good_out_list)
	debug(good_out_list)

def collect_long_execute_cmd(tn,cmd,**kwargs):
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 8
	if "prompt" in kwargs:
		prompt = kwargs['prompt']
	else:
		prompt = "# "
	#relogin_if_needed(tn)
	original_cmd = cmd
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(cmd_bytes)
	output = switch_read_console_output(tn,timeout = timeout,prompt = prompt)
	good_out_list = clean_show_output_recursive(output,original_cmd)
	print(good_out_list)
	return good_out_list

def ftg_collect_execute_cmd(tn,cmd,**kwargs):
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 8
	#relogin_if_needed(tn)
	original_cmd = cmd
	tprint(f"Executing command: {cmd}")
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(cmd_bytes)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	sleep(timeout)
	output = tn.read_very_eager()
	dprint(output)
	#output = tn.read_until(("# ").encode('ascii'))
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)
	
	dprint(out_str_list)
	return out_str_list

def collect_execute_cmd(tn,cmd,**kwargs):
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 8
	#relogin_if_needed(tn)
	original_cmd = cmd
	cmd_bytes = convert_cmd_ascii_n(cmd)
	tn.write(cmd_bytes)
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	tn.write(('' + '\n').encode('ascii')) # uncomment this line if doesn't work
	sleep(timeout)
	output = tn.read_very_eager()
	dprint(output)
	#output = tn.read_until(("# ").encode('ascii'))
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).strip(' \r')
		out_str_list.append(o_str)
	 
	if cmd == "get router info6 bgp summary":
		print (f"return from utiliy.py: collect_show_cmd(): {out_str_list}")

	good_out_list = clean_show_output_recursive(out_str_list,original_cmd)
	dprint(good_out_list)
	return good_out_list

def process_show_command(output):
	out_list = output.split('\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).rstrip(' ')
		out_str_list.append(o_str)
	if cmd == "get router info6 bgp summary":
		print (f"return from utiliy.py: collect_show_cmd(): {out_str_list}")
	return out_str_list

def append_file_collect_show(filename,result):
	singleline = "-------------------------------------------------------------------------------------\n"
	with open(filename,'a+') as f:
		f.write(singleline)
		for line in result:
			f.write(time_str(f'{line}\n'))
			 

def collect_show_cmd_fast(tn,cmd,**kwargs):
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 2
	#relogin_if_needed(tn)
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	sleep(1)
	output = tn.read_very_eager()
	#output = tn.read_until(("# ").encode('ascii'))
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).rstrip(' ')
		out_str_list.append(o_str)
	# tprint(dir(output))
	# tprint(type(output))
	#tprint(out_list)
	# for i in out_str_list:
	# 	tprint(i)
	i = 0
	for o in out_str_list:
		if cmd in o:
			break
		out_str_list.remove(o)
	return out_str_list

def get_mac_table_size(dut):
	cmd = "diagnose switch mac-address list | grep MAC | wc -c"
	result = switch_show_cmd(dut,cmd,t=3)
	debug(result)
	r = ''
	for r in result:
		if "CLI" in r:
			break
	debug(f'r = {r}')
	if r=='':
		tprint("Error getting switch MAC table size")
		return 0
	try:
		s = r.split(":")
		size = int(s[1])
	except Exception as e:
		tprint("Error getting switch MAC table size")
		return 0
	return size

def switch_exec_cmd(tn,cmd,**kwargs):
	#relogin_if_needed(tn)
	if 't' in kwargs:
		wait = kwargs['t']
	else:
		wait = 10
	tprint(f"Executing command: {cmd}")
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	tn.read_until(("# ").encode('ascii'),timeout=wait)
	

def switch_run_cmd(tn,cmd,**kwargs):
	#relogin_if_needed(tn)
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 0
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	sleep(timeout)
	 
def switch_show_cmd_name(dut_dir,cmd,**kwargs):
	#relogin_if_needed(tn)
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 2
	tn = dut_dir['telnet']
	name = dut_dir['name']
	old_cmd = cmd
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	sleep(timeout)
	output = tn.read_very_eager()
	#output = tn.read_until(("# ").encode('ascii'))
	out_list = output.split(b'\r\n')
	debug(f"out_list in switch_show_cmd = {out_list}")
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		try:
			o_str = o.decode(encoding).rstrip(' ')
			out_str_list.append(o_str)
		except Exception as e:
			pass
	# tprint(dir(output))
	# tprint(type(output))
	#tprint(out_list)
	tprint(f"----------------{name}: {old_cmd} ---------------")
	for i in out_str_list:
		tprint(i)
	return out_str_list

def print_collect_show(output):
	for i in output:
		tprint(i)

def switch_show_cmd_linux(tn,cmd,**kwargs):
	#relogin_if_needed(tn)
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 2
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	tn.write(('' + '\n').encode('ascii'))
	# tn.write(('' + '\n').encode('ascii'))
	# tn.write(('' + '\n').encode('ascii'))
	# tn.write(('' + '\n').encode('ascii'))
	sleep(timeout)
	output = tn.read_very_eager()
	#output = tn.read_until(("# ").encode('ascii'))
	out_list = output.split(b'\r\n')
	debug(f"out_list in switch_show_cmd = {out_list}")
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		try:
			o_str = o.decode(encoding).rstrip(' ')
			out_str_list.append(o_str)
		except Exception as e:
			pass
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	# tprint(dir(output))
	# tprint(type(output))
	#tprint(out_list)
	for i in out_str_list:
		tprint(i)
	return out_str_list

def increment_mac_address(*args,**kwargs):
	start_mac = kwargs['start_mac']
	num = kwargs['num']

	#mac="0xaabbccdd0000"
	mac = "0x"+start_mac.replace(":","")
	mac_addresses = []
	for i in range(num):
	    mac = "{:012X}".format(int(mac, 16) + 1)
	    new_mac = (':'.join(mac[i]+mac[i+1] for i in range(0, len(mac), 2)))
	    mac_addresses.append(new_mac)
	print(mac_addresses)
	return mac_addresses


def switch_show_cmd(tn,cmd,**kwargs):
	#relogin_if_needed(tn)
	if 't' in kwargs:
		timeout = kwargs['t']
	else:
		timeout = 2
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	sleep(timeout)
	output = tn.read_very_eager()
	#output = tn.read_until(("# ").encode('ascii'))
	out_list = output.split(b'\r\n')
	debug(f"out_list in switch_show_cmd = {out_list}")
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		try:
			o_str = o.decode(encoding).rstrip(' ')
			out_str_list.append(o_str)
		except Exception as e:
			pass
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	#tprint(dir(output))
	# tprint(type(output))
	#tprint(out_list)
	for i in out_str_list:
		tprint(i)
	return out_str_list
###############################################################################
#   config_keys =  "configure switch interface \
#						|  edit port1 \
#						|  set status up \
#						|  end \
#				"
###############################################################################

def ist_add(proc_list_all,proc_list):

	for i in proc_list1:
		proc_list_all.append(i)


def press_any_key():
	print_dash_line()
	keyin = input(f"Press any key to continue...")

def config_cmds_lines_cisco(dut, cmdblock):
	b= cmdblock.split("\n")
	b = [x.strip() for x in b if x.strip()]
	for cmd in b:
		switch_configure_cmd_cisco(dut,cmd)

def split_f_string_lines(cmdblock):
	b= cmdblock.split("\n")
	b = [x.strip() for x in b if x.strip()]
	return b

def split_fstring_lines_generic(cmdblock):
	cmds= cmdblock.split("\n")
	print(cmds)
	while cmds and cmds[-1] == '':
	    cmds.pop()
	while cmds and cmds[0] == '':
	    cmds.pop(0)
	return cmds


def config_cmds_lines_fast(dut,cmdblock,*args,**kwargs):
	if "wait" in kwargs:
		wait_time = int(kwargs["wait"])
	else:
		wait_time = 0.2

	if "feedback" in kwargs:
		feedback = kwargs['feedback']
	else:
		feedback = False

	if "check_prompt" in kwargs:
		check_prompt = kwargs["check_prompt"]
	else:
		check_prompt = False

	b= cmdblock.split("\n")
	b = [x.strip() for x in b if x.strip()]
	config_return_list = []
	for cmd in b:
		config_return = switch_configure_cmd(dut,cmd,output=feedback,mode="fast")
		if config_return != None:
			config_return_list.append(config_return)
		sleep(wait_time)
	return config_return_list

def config_cmds_lines(dut,cmdblock,*args,**kwargs):
	
	if "wait" in kwargs:
		wait_time = int(kwargs["wait"])
	else:
		wait_time = 0.2

	if "feedback" in kwargs:
		feedback = kwargs['feedback']
	else:
		feedback = False

	if "check_prompt" in kwargs:
		check_prompt = kwargs["check_prompt"]
	else:
		check_prompt = False


	if "mode" in kwargs:
		config_mode = kwargs["mode"]
	elif "device" in kwargs:
		device= kwargs['device']
		current_time = time.time()
		if device.last_cmd_time == None:
			config_mode = "fast"
		elif (current_time - device.last_cmd_time) < 100:
			config_mode = "fast"
		else:
			config_mode = "slow"
		device.last_cmd_time = current_time
	else:
		config_mode = "slow"

	if config_mode == "fast":
		wait_time = 0.5
		check_prompt = False
	else:
		check_prompt = True

	if check_prompt:
		handle_prompt_before_commands(dut)

	b= cmdblock.split("\n")
	b = [x.strip() for x in b if x.strip()]
	config_return_list = []
	for cmd in b:
		config_return = switch_configure_cmd(dut,cmd,mode=config_mode,output=feedback)
		if config_return != None:
			config_return_list.append(config_return)
		sleep(wait_time)

	return config_return_list
		 
def print_attributes(fgt):
	attrs = vars(fgt)
	print(attrs)
	print_dict_simple(attrs)

def print_dict_simple(attrs):
	for k,v in attrs.items():
		print(f"{k}:{v}")

def config_block_cmds_new(dut, cmdblock):
	b= cmdblock.split("\n")
	b = [x.strip() for x in b if x.strip()]
	for cmd in b:
		switch_configure_cmd(dut,cmd)

def print_title(msg):
	print(f"================================ {msg} ===============================")

def config_block_cmds(dut_dir, cmdblock):
	b= cmdblock.split("\n")
	b = [x.strip() for x in b if x.strip()]
	for cmd in b:
		switch_configure_cmd_name(dut_dir,cmd)


def switch_configure_cmd_name(dut_dir,cmd):
	tn = dut_dir['telnet']
	dut_name = dut_dir['name']
	# swn9k-1: config t 2 :: configuring
	# swn9k-1: line con 2 :: configuring
	# swn9k-1(config): exec-timeout 300
	# swn9k-1(config-console): end uring
	tprint(f"configuring {dut_name}: {cmd}")
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	time.sleep(0.5)
	tn.read_until(("# ").encode('ascii'),timeout=10)

def switch_read_console_output(tn,**kwargs):
	if "timeout" in kwargs:
		t = kwargs['timeout']
	else:
		t = 10
	if "prompt" in kwargs:
		prompt = kwargs['prompt']
	else:
		prompt = "# "
	output = tn.read_until((prompt).encode('ascii'),timeout=t)
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	out_str_list = []
	for o in out_list:
		o_str = o.decode(encoding).rstrip(' ')
		out_str_list.append(o_str)
	# tprint(dir(output))
	# tprint(type(output))
	# tprint(out_list)
	for i in out_str_list:
		tprint(i)
	return out_str_list

# def switch_config_cmd_dutinfo(original_func):
# 	def wrapper(*args,**kwargs):


def switch_configure_cmd(tn,cmd,**kwargs):
	if 'mode' in kwargs:
		mode = kwargs['mode']
	else:
		mode = None
	if 'output' in kwargs:
		output = kwargs["output"]
	else:
		output = False

	if mode == "fast":
		tprint(f"configuring {cmd}")
	else:
		dut_prompt = find_dut_prompt(tn)
		tprint("configuring {}: {}".format(dut_prompt,cmd))

	#cmd = convert_cmd_ascii_n(cmd)
	enter_console_cmd(tn,cmd)
	time.sleep(0.6)
	if output == False:
		tn.read_until(("# ").encode('ascii'),timeout=5)
		return None 
	else: 
		#sleep(0.2)
		config_output = tn.read_very_eager()
		out_list = config_output.split(b'\r\n')
		encoding = 'utf-8'
		out_str_list = []
		for o in out_list:
			o_str = o.decode(encoding).strip(' \r')
			out_str_list.append(o_str)
		 
		for n in out_str_list:
			print (f"{n}")
		return out_str_list


def telnet_send_cmd(tn,cmd,*args,**kwargs):
	cmd = convert_cmd_ascii_n(cmd) #convert_cmd_ascii_n has appended return at the end
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	tn.read_until((">").encode('ascii'),timeout=2)
	tn.write(cmd)
	time.sleep(2)
	tn.read_until((">").encode('ascii'),timeout=2)
	#tn.expect(["#", ">","> ",">  ",">	"])

def switch_configure_cmd_cisco(tn,cmd,**kwargs):
	if 'mode' in kwargs:
		mode = kwargs['mode']
	else:
		mode = None

	if mode == "silent":
		pass
	else:
		dut_prompt = find_dut_prompt_cisco(tn)
	# swn9k-1: config t 2 :: configuring
	# swn9k-1: line con 2 :: configuring
	# swn9k-1(config): exec-timeout 300
	# swn9k-1(config-console): end uring
		tprint("configuring {}: {}".format(dut_prompt,cmd))
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	time.sleep(0.2)
	tn.read_until(("# ").encode('ascii'),timeout=10)

def switch_wait_enter_yes(tn,prompt):
	prompt = convert_cmd_ascii(prompt)
	#prompt_re = (prompt + r'.*').encode('ascii')
	tn.read_until(prompt,timeout=30)
	time.sleep(1)

	answer = convert_cmd_ascii('y' )
	tn.write(answer)
	time.sleep(1)
	 
def switch_enter_yes(tn):
	# prompt = convert_cmd_ascii(prompt)
	# #prompt_re = (prompt + r'.*').encode('ascii')
	# tn.read_until(prompt,timeout=60)
	# time.sleep(1)

	answer = convert_cmd_ascii('y')
	tn.write(answer)
	time.sleep(2)

def switch_interactive_exec_bios(tn,exec_cmd,prompt):
	#relogin_if_needed(tn)
	tprint(exec_cmd)
	exec_cmd = exec_cmd
	exec_cmd = convert_cmd_ascii(exec_cmd)
	tn.write(exec_cmd)
	time.sleep(1)

	answer = convert_cmd_ascii('y')
	#answer = convert_cmd_ascii('y' + '\n')
	tn.write(answer)
	time.sleep(1)

def switch_interactive_exec(tn,exec_cmd,prompt):
	#relogin_if_needed(tn)
	tprint(exec_cmd)
	exec_cmd = exec_cmd + '\n'
	exec_cmd = convert_cmd_ascii(exec_cmd)
	tn.write(exec_cmd)
	output = tn.read_very_eager()
	time.sleep(1)

	prompt = convert_cmd_ascii(prompt)
	#prompt_re = (prompt + r'.*').encode('ascii')
	tn.read_until(prompt,timeout=10)
	time.sleep(1)

	answer = convert_cmd_ascii('y')
	#answer = convert_cmd_ascii('y' + '\n')
	tn.write(answer)
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	for i in range(5):
		tn.read_very_eager()
	time.sleep(1)


def find_dut_prompt_cisco(tn):
	tn.write(('' + '\n').encode('ascii'))
	#print("Reading prompt and retrieve prompt......")
	output = tn.read_until(("#").encode('ascii'))
	out_list = output.split(b'\r\n')
	encoding = 'utf-8'
	for o in out_list:
		o_str = o.decode(encoding).rstrip(' ')
		if "#" in o_str:
			prompt = o_str.strip(' ')
			prompt = prompt.strip("#")
	#print(prompt)
	return prompt



def reliable_telnet(ip_address,*args,**kwargs):
	if 'sig' in kwargs:
		event = kwargs['sig']
		end = event.is_set()
	else:
		end = False
	while not end:
		handle = telnet_connection(ip_address)
		if handle == False:
			sleep(3)
			if 'sig' in kwargs:
				event = kwargs['sig']
				end = event.is_set()
			else:
				end = False
			continue
		else:
			return handle

def telnet_connection(ip_address,**kwargs):
	tprint(f"Device management interface = {str(ip_address)}")
	if "password" in kwargs:
		pwd = kwargs["password"]
	else:
		pwd = 'xxxx123'
	user = 'admin'
	#switch_login(ip_address,console_port)
	try:
		tn = telnetlib.Telnet(ip_address,23,10)
	except Exception as e: 
		tprint("!!!!!!!!!!!Telnet is either time out or not response from device, Need to retry later")
		# sleep(2)
		# tn = telnetlib.Telnet(ip_address,console_port_int)
		return False

	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	tn.read_until(("login: ").encode('ascii'),timeout=5)
	tn.write((user + '\n').encode('ascii'))
	tn.read_until(("Password: ").encode('ascii'),timeout=5)
	tn.write((pwd + '\n').encode('ascii'))
	tn.read_until(("# ").encode('ascii'),timeout=5)
	tn.write(('' + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	return tn

def gnetch_command(cmd,target):
	gnetch_cmd = f"stubby --proto2 call blade:gnetch-frontend Gnetch2.Command 'command: \"{cmd}\" target: \"{target}\"'"
	output = execute_os_cmd(gnetch_cmd)
	output = output[0].lstrip("data: ").strip('"').split("\\n")
	return output[:-1]
	#return output[1:-1]


def execute_os_cmd(command):
    result = subprocess.run(command, shell=True, capture_output=True, text=True)
    # Split the output into a list of lines
    output_lines = result.stdout.splitlines()
    return output_lines

def test_ping_window(ip_address,verify=True):

  def analyze_ping_output(ping_output):
      packet_unreachable_match = re.search(r' unreachable', ping_output)
      if packet_unreachable_match:
          return False,"Destination host unreachable"

      packet_loss_match = re.search(r'Lost = (\d+) \((\d+)% loss\)', ping_output)
      
      if packet_loss_match:
          lost_packets = int(packet_loss_match.group(1))
          loss_percentage = int(packet_loss_match.group(2))
          if lost_packets == 0:
              return True, "Ping successful"
          else:
              return False, f"Packet loss: {loss_percentage}%"
      else:
          return False, "Packet loss information not found"

  def run_ping(ip_address):
      try:
          ping_process = subprocess.Popen(['ping', ip_address], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
          ping_output, _ = ping_process.communicate()
          return ping_output
      except Exception as e:
          return str(e)

  ping_output = run_ping(ip_address)
  print("Ping Output:")
  print(ping_output)
  if verify == True:
	  is_good, result_message = analyze_ping_output(ping_output)

	  if is_good:
	      print(f"Ping {ip_address} is good: {result_message}")
	      return True
	  else:
	      print(f"Ping {ip_address} is not successful: {result_message}")
	      return False

def ping_ipv4(tn,*args,**kwargs):
	ip = kwargs["ip"]
	result = collect_execute_cmd(tn,f"execute ping {ip}")
	dprint(result)
	print_cmd_output_from_list(result)

def ping_ipv6(tn,*args,**kwargs):
	ip = kwargs["ip"]
	result = collect_execute_cmd(tn,f"execute ping6 {ip}")
	dprint(result)
	print_cmd_output_from_list(result)

def ping_ipv6_extensive(*args,**kwargs):
	#ping_ipv6_extensive(console=self.switches[i].console, ip_src=ip_src,ip_dst = ip_dst,name=interface_name)
	tn = kwargs["console"]
	ip_src = kwargs["ip_src"]
	ip_dst = kwargs["ip_dst"]
	interface_name = kwargs['name']
	sw_src = kwargs['sw_src']
	sw_dst = kwargs['sw_dst']
	Info(f"=================== Ping source: {interface_name}: 2nd IPv6 {ip_src} on {sw_src} ===============")
	show_execute_cmd(tn,f"execute ping6-options source {ip_src}")
	Info(f"=================== Ping Destination: {ip_dst} on {sw_dst} ===============")
	show_execute_cmd(tn,f"execute ping6 {ip_dst}")
 

def telnet_device_universal(ip_address,port=23,*args,**kwargs):
	TIMEOUT = 5
	tprint("Device IP Address ="+str(ip_address))
	tprint("Telnet Port ="+str(port))
	if "password" in kwargs:
		pwd = kwargs["password"]
	else:
		pwd = 'ciena123'
	if "platform" in kwargs:
		platform = kwargs['platform']
	else:
		platform = "ciena"

	if "login" in kwargs:
		login = kwargs["login"]
	else:
		login = "diag"

	port_int = int(port)
 
	try:
		tn = telnetlib.Telnet(ip_address,port_int,5)
	except ConnectionRefusedError: 
		tprint(f"!!!!!!!!!!! the host {ip_address} is refused to connect !!!!!!")
		tn = None 
		return tn
	except TimeoutError: 
		tprint(f"!!!!!!!!!!!Telnet to {ip_address} timeout")
		tn = None 
		return tn
	except socket.error:
		tprint("Socket error")
		reason = "socket"
		return None
	except Exception as e:
		tprint(f"!!!!!!!!!!! Something is wrong with telnet {ip_address}, check your network connectitiy")
		tn = None 
		return tn
 
	tn.read_until(("login: ").encode('ascii'),timeout=5)
	tn.write((login + '\n').encode('ascii'))
	tn.read_until(("Password: ").encode('ascii'),timeout=5)
	tn.write((pwd + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	sleep(2)
	tn.write(('' + '\n').encode('ascii'))
	login_result = tn.read_very_eager()
	Errors = ["incorrect","Invalid"]
	for err in Errors:
		if err in login_result.decode():
			print(f"Login into {ip_address} is not successful, please check login/password")
			return None
	return tn

def telnet_device_ciena(ip_address,port=23,*args,**kwargs):
	TIMEOUT = 5
	tprint("Device IP Address ="+str(ip_address))
	tprint("Telnet Port ="+str(port))
	if "password" in kwargs:
		pwd = kwargs["password"]
	else:
		pwd = 'ciena123'
	if "platform" in kwargs:
		platform = kwargs['platform']
	else:
		platform = "ciena"

	if "login" in kwargs:
		login = kwargs["login"]
	else:
		login = "diag"

	port_int = int(port)
 
	try:
		tn = telnetlib.Telnet(ip_address,port_int,5)
	except ConnectionRefusedError: 
		tprint(f"!!!!!!!!!!! the host {ip_address} is refused to connect !!!!!!")
		tn = None 
		reason = "refuse"
		return tn,reason
	except TimeoutError: 
		tprint(f"!!!!!!!!!!!Telnet to {ip_address} timeout")
		tn = None 
		reason = "timeout"
		return tn,reason
	except socket.error:
		tprint("Socket error")
		reason = "socket"
		return tn,reason
	except Exception as e:
		tprint(f"!!!!!!!!!!! Something is wrong with telnet {ip_address}, check your network connectitiy")
		tn = None 
		reason = "unknown"
		return tn,reason
 
	tn.read_until(("login: ").encode('ascii'),timeout=5)
	tn.write((login + '\n').encode('ascii'))
	tn.read_until(("Password: ").encode('ascii'),timeout=5)
	tn.write((pwd + '\n').encode('ascii'))
	tn.write(('' + '\n').encode('ascii'))
	sleep(2)
	tn.write(('' + '\n').encode('ascii'))
	login_result = tn.read_very_eager()
	Errors = ["incorrect","Invalid"]
	for err in Errors:
		if err in login_result.decode():
			print(f"Login into {ip_address} is not successful, please check login/password")
			reason = "password"
			tn = None
			return tn,reason
	reason = "success"
	return tn,reason

def file_exists(file_path):
    if os.path.exists(file_path):
        return True 
    else:
        return False

def is_valid_ip(ip):
    try:
        ip_network = ipaddress.IPv4Network(ip, strict=False)
        return True
    except ValueError:
        return False

def decrement_ip_address(ip):
	ip_object = ipaddress.IPv4Address(ip)
	decremented_ip_object = ip_object - 1
	decremented_ip = str(decremented_ip_object)
	return decremented_ip

def log_show_cmd_wexpect(handle,cmd,t=2):
	output, error = wexpect_send_cmd_expect_prompt(handle,cmd,wait=t)
	out_list = output.split('\r\n')
	encoding = 'utf-8'
	out_str_list = []
	empty_list = []
	for o in out_list:
		o_str = o.strip(' \r')
		out_str_list.append(o_str)

	good_out_list = clean_show_output_recursive(out_str_list,cmd)
	if good_out_list == None:
		return empty_list
	try: 
		for msg in good_out_list:
			print(msg)
		return good_out_list
	except Exception as e:
		return empty_list

def log_show_cmd_pexpect(handle,cmd,t=2):
	output, error = wexpect_send_cmd_expect_prompt(handle,cmd,wait=t)
	out_list = output.decode('utf-8').split('\r\n')
	encoding = 'utf-8'
	out_str_list = []
	empty_list = []
	for o in out_list:
		o_str = o.strip(' \r')
		out_str_list.append(o_str)

	good_out_list = clean_show_output_recursive(out_str_list,cmd)
	if good_out_list == None:
		return empty_list
	try: 
		for msg in good_out_list:
			print(msg)
		return good_out_list
	except Exception as e:
		return empty_list


def wexpect_send_simple_cmd(child,cmd):
    dprint(f"CLI command: {cmd}")
    child.sendline(cmd)

def wexpect_send_cmd_expect_prompt(child,cmd,wait=2):
	dprint(f"wait time = {wait}")
	dprint(f"CLI command: {cmd}")
	#cmd = convert_cmd_ascii_n(cmd)
	wexpect_send_simple_cmd(child,cmd)
	prompts = ['[A-Za-z0-9\"\']+>','[\"\'A-Za-z0-9]+#']
	child.expect(prompts,timeout=wait)
	result = child.before
	error = child.after
	dprint(f"Result of command: {cmd}\n{result}")
	dprint(f"After command: {cmd}\n{error}")
	return(result,error)

def pexpect_ssh_sessions(ip,username="ADMIN",password="ADMIN",port=22):
	real_executable = sys.executable
	try:
	    if sys._MEIPASS is not None:
	        sys.executable = os.path.join(sys._MEIPASS, "wexpect", "wexpect.exe")
	except AttributeError:
	    pass
	try:
		ssh_port = port
		ssh_cmd = f'ssh -o ConnectTimeout=10 {username}@{ip} -p {ssh_port}'
		print(f"ssh_cmd = {ssh_cmd}")
		child =pexpect.spawn(ssh_cmd,timeout=10)
		sys.executable = real_executable
	except (ImportError, OSError, IOError,ValueError):
		print("IO error")
	#the time out is needed to set to avoid child expect to time out,don't set to 5000
	child.timeout = 1000
	child.maxread = 10000000
	ssh_prompt = "Are you sure you want to continue connecting (yes/no/[fingerprint])?"
	i = child.expect(['Login:', 'fingerprint','Connection closed',pexpect.EOF, pexpect.TIMEOUT])

	# If the prompt is for the fingerprint, send "yes\r"
	if i == 0:
		child.sendline(username)
	elif i == 1:
		child.sendline('yes\r')
		child.expect(['Login:'])
		child.sendline(username)
	elif i == 2:
		print(f"Connection to {ip} was closed by remote host")
		return None 
	elif i == 3:
		print(f"Connection to {ip} error EOF")
		return None 
	elif i == 4:
		print(f"Connection to {ip} Time out")
		return None 
	child.expect("Password:") 
	child.sendline(password)
	prompts = ['#']
	child.expect(prompts,timeout=2)
	result = child.before
	error = child.after
	dprint(f"Result of ssh: {ssh_cmd}\n{result}")
	dprint(f"After ssh: {ssh_cmd}\n{error}")
	return child


def wexpect_ssh_sessions(ip,username="ADMIN",password="ADMIN",port=22):
	real_executable = sys.executable
	try:
	    if sys._MEIPASS is not None:
	        sys.executable = os.path.join(sys._MEIPASS, "wexpect", "wexpect.exe")
	except AttributeError:
	    pass
	try:
		ssh_port = port
		ssh_cmd = f'ssh {username}@{ip} -p {ssh_port}'
		print(f"ssh_cmd = {ssh_cmd}")
		child =wexpect.spawn(ssh_cmd)
		sys.executable = real_executable
	except (ImportError, OSError, IOError,ValueError):
		print("IO error")
	#the time out is needed to set to avoid child expect to time out,don't set to 5000
	child.timeout = 1000
	child.maxread = 10000000
	ssh_prompt = "Are you sure you want to continue connecting (yes/no/[fingerprint])?"
	i = child.expect(['Login:', 'fingerprint','Connection closed',wexpect.EOF, wexpect.TIMEOUT])

	# If the prompt is for the fingerprint, send "yes\r"
	if i == 0:
		child.sendline(username)
	elif i == 1:
		child.sendline('yes\r')
		child.expect(['Login:'])
		child.sendline(username)
	elif i == 2:
		print(f"Connection to {ip} was closed by remote host")
		return None 
	elif i == 3:
		print(f"Connection to {ip} error EOF")
		return None 
	elif i == 4:
		print(f"Connection to {ip} Time out")
		return None 
	child.expect("Password:") 
	child.sendline(password)
	prompts = ['#']
	child.expect(prompts,timeout=2)
	result = child.before
	error = child.after
	dprint(f"Result of ssh: {ssh_cmd}\n{result}")
	dprint(f"After ssh: {ssh_cmd}\n{error}")
	return child


def telnet_switch(ip_address, console_port,*args,**kwargs):
	TIMEOUT = 5
	tprint("console server ip ="+str(ip_address))
	tprint("console port="+str(console_port))
	if "password" in kwargs:
		pwd = kwargs["password"]
	else:
		pwd = 'ciena123'
	if "platform" in kwargs:
		platform = kwargs['platform']
	else:
		platform = "ciena"

	if "login" in kwargs:
		login = kwargs["login"]
	else:
		login = "diag"

	
	console_port_int = int(console_port)
 
	try:
		tn = telnetlib.Telnet(ip_address,console_port_int)
	except ConnectionRefusedError: 
		tprint("!!!!!!!!!!!the console is being used, need to clear it first")
		tn = None 
		return tn
	except TimeoutError: 
		tprint("!!!!!!!!!!!Telnet to console server timeout, skip this switch console connection")
		tn = None 
		return tn
	except Exception as e:
		tprint("!!!!!!!!!!! Something is wrong with telnet connection, check your network connectitiy to console server")
		tn = None 
		return tn
	 
	for i in range(6):
		tn.write(('' + '\n').encode('ascii'))
		sleep(0.2)
	 
 
	debug("See what prompt the console is at")
	output = tn.expect([re.compile(b"login:")],timeout=TIMEOUT)
	debug(f"collect info to look for prompt: {output}")
	debug(f"output[2] = {output[2].decode().strip()}")
	#debug(output[2].decode().strip())
	prompt = output[2].decode().strip()
 
	if 'diag' in prompt:
		debug (f"It is in diag shell, need to exit")
		tn.write(("exit" + '\n').encode('ascii'))
		tn.read_until((">").encode('ascii'),timeout=5)
		tn.write(('' + '\n').encode('ascii'))
	elif output[0] == 0:
		debug("it is a login prompt, you need to re-login")
		tn.read_until(("login: ").encode('ascii'),timeout=5)
		tn.write((login + '\n').encode('ascii'))
		tn.read_until(("Password: ").encode('ascii'),timeout=5)
		tn.write((pwd + '\n').encode('ascii'))
		tn.write(('' + '\n').encode('ascii'))
		sleep(1)
		tn.write(('' + '\n').encode('ascii'))
	elif ">" in prompt: 
		debug("it is a Shell exect prompt")
		pattern = r"[a-zA-Z0-9\-]+"
		match = re.match(pattern,prompt)
		if match:
			result = match.group()
			Info(result)
	elif "#" in prompt: 
		debug("it is a config mode prompt")
		pattern = r"[a-zA-Z0-9\-]+"
		match = re.match(pattern,prompt)
		if match:
			result = match.group()
			debug(result)
		tn.write(("exit" + '\n').encode('ascii'))
		tn.read_until((">").encode('ascii'),timeout=5)
	else:
		Info("can not get any prompt, it is very likely the switch is not reachable ...")
		tn = None
	if tn == None:
		return tn

	tn.write(('set session more off' + '\n').encode('ascii'))
	result = collect_show_cmd_ciena(tn,"show software",t=4)
	return tn
										   

def console_timer(seconds,**kwargs):
	if 'msg' in kwargs:
		notice = kwargs['msg']
		tprint(f'========================= {notice} ==========================')
	for remaining in range(seconds, 0, -1):
		sys.stdout.write("\r")
		sys.stdout.write("============================ Timer:{:2d} seconds remaining =======================".format(remaining))
		sys.stdout.flush()
		time.sleep(1)
	sys.stdout.write("\n")

	if switch_find_login_prompt(tn) == True:
		# tn.read_until(("login: ").encode('ascii'),timeout=10)
		tn.write(('admin' + '\n').encode('ascii'))
		tn.read_until(("Password: ").encode('ascii'),timeout=10)
		tn.write((password + '\n').encode('ascii'))
		sleep(1)
		tn.read_until(("# ").encode('ascii'),timeout=10)
		switch_configure_cmd(tn,'config system global',mode="silent")
		switch_configure_cmd(tn,'set admintimeout 480',mode="silent")
		switch_configure_cmd(tn,'end',mode="silent")
		return True
	else:
		switch_configure_cmd(tn,'config system global',mode='silent')
		switch_configure_cmd(tn,'set admintimeout 480',mode='silent')
		switch_configure_cmd(tn,'end',mode='silent')
		return 

def relogin_if_needed(tn):
	if settings.TELNET:
		try:
			tn.write(('' + '\n').encode('ascii'))
			sleep(2)
			tn.write(('' + '\n').encode('ascii'))
			sleep(2)
			tn.write(('' + '\n').encode('ascii'))
			sleep(2)
			tn.write(('' + '\n').encode('ascii'))
			sleep(2)
			tn.write(('' + '\n').encode('ascii'))
			sleep(2)
		except BrokenPipeError:
			return False

	if switch_find_login_prompt(tn) == True:
		switch_login(tn,mode='silent')
		return True
	else:
		switch_configure_cmd(tn,'config system global',mode='silent')
		switch_configure_cmd(tn,'set admintimeout 480',mode='silent')
		switch_configure_cmd(tn,'end',mode='silent')
		return 
		

def find_shell_prompt(tn,chassis_id):
	TIMEOUT = 5
	out = tn.expect([re.compile(b"#")],timeout=TIMEOUT)
	dprint(f"after enter password, device prompt overall prompt = {out}")
	login_result = out[0]
	device_prompt = out[2].decode().strip()
	dprint(f"Expecting # prompt, if return 0, # is found. return = {login_result},device prompt ={device_prompt}")
	if int(login_result) == 0 and chassis_id in device_prompt:
		tprint(f"login successful to {chassis_id}!")
		return True
	else:
		return False

def ciena_config_cmd(tn,cmd):
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	tn.read_until(("#").encode('ascii'),timeout=5)

def ciena_exec_cmd(tn,cmd):
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	tn.read_until((">").encode('ascii'),timeout=5)

def enter_console_cmd(tn,cmd):
	cmd = convert_cmd_ascii_n(cmd)
	tn.write(cmd)
	gabage = tn.read_very_eager()
	dprint(gabage)

def clear_console_buffer(tn):
	for i in range(10):
		tn.write(('' + '\n').encode('ascii'))
		sleep(0.5)
		tn.read_very_eager()

def increment_32(ip,num):
	bytes = ip.split('.')
	ibytes = [int(i) for i in bytes]
	newip_list = [ip]
	for i in range(num-1):
		ibytes[3] += 1
		if ibytes[3] > 255:
			ibytes[2] += 1
			ibytes[3] = 0
			if ibytes[2] > 255:
				ibytes[1]+= 1
				ibytes[2] = 0
				if ibytes[1] > 255:
					ibytes[0]+= 1
					ibytes[1] = 0
					if ibytes[0] > 224:
						print("The range is too big for IPv4")
						return newip_list
		newip = ".".join(str(i) for i in ibytes)
		newip_list.append(newip)
	print(newip_list)
	return newip_list


def increment_24(ip,num):
	bytes = ip.split('.')
	ibytes = [int(i) for i in bytes]
	newip_list = [ip]
	for i in range(num-1):
		ibytes[2] += 1
		if ibytes[2] > 255:
			ibytes[1] += 1
			ibytes[2] = 0
			if ibytes[1] > 255:
				ibytes[0]+= 1
				ibytes[1] = 0
				if ibytes[0] > 224:
					print("The range is too big for IPv4")
					return newip_list
		newip = ".".join(str(i) for i in ibytes)
		newip_list.append(newip)
	print(newip_list)
	return newip_list

def handle_prompt_before_commands(tn,*args,**kwargs):
	Info("Before entering commands into device, find out what prompt the device is at")
	password = "ciena123"
	tn.write(('' + '\n').encode('ascii'))
	sleep(0.5)
	tn.write(('' + '\n').encode('ascii'))
	sleep(0.5)
	tn.write(('' + '\n').encode('ascii'))
	sleep(0.5)
	tn.write(('' + '\n').encode('ascii'))
	sleep(0.5)
	tn.write(('' + '\n').encode('ascii'))
	sleep(0.5)
	tn.write(('' + '\n').encode('ascii'))
	sleep(0.5)
	TIMEOUT = 3
	Info("See what prompt the console is at")
	output = tn.expect([re.compile(b"login:")],timeout=TIMEOUT)
	Info(f"collect info to look for prompt: {output}")
	Info(f"output[2] = {output[2].decode().strip()}")
	#debug(output[2].decode().strip())
	prompt = output[2].decode().strip()
	if output[0] == 0:
		Info("it is a login prompt, you need to re-login because of timeout or reboot")
		tn.write(('' + '\n').encode('ascii'))
		tn.write(('' + '\n').encode('ascii'))
		tn.write(('' + '\n').encode('ascii'))
		tn.write(('' + '\n').encode('ascii'))
		sleep(1)
		tn.read_until(("login: ").encode('ascii'),timeout=10)
		tn.write(('diag' + '\n').encode('ascii'))           # this would not work for factory reset scenario
		tn.read_until(("Password: ").encode('ascii'),timeout=10)
		tn.write((password + '\n').encode('ascii'))
		tn.write(('' + '\n').encode('ascii'))
		tn.write(('' + '\n').encode('ascii'))
		sleep(0.2)
		tn.read_until(("# ").encode('ascii'),timeout=10)
		return ("re-login",None)
	elif ">" in prompt or "#" in prompt: # be careful of with and without space at the front
		Info("it is a Shell prompt")
		pattern = r"[a-zA-Z0-9\-]+"
		match = re.match(pattern,prompt)
		if match:
			result = match.group()
		else:
			result = None
		return ("shell",result)
	else:
		debug("can not get any prompt, need to use robust login procedure...")
		return (None,None)


def switch_find_login_prompt(tn):
	TIMEOUT = 4
	# tn.write(('\x03').encode('ascii'))
	# tn.write(('\x03').encode('ascii'))
	tn.write(('\x03').encode('ascii'))
	time.sleep(2)
	tn.write(('' + '\n').encode('ascii'))
	time.sleep(1)
	tn.write(('' + '\n').encode('ascii'))
	time.sleep(1)
	tn.write(('' + '\n').encode('ascii'))
	time.sleep(1)
	tn.write(('' + '\n').encode('ascii'))
	time.sleep(1)
	#time.sleep(1)
	
	debug("See what prompt the console is at")
	output = tn.expect([re.compile(b"login:")],timeout=TIMEOUT)
	debug(output[2].decode().strip())
	if output[0] < 0: 
		debug("It is a NOT login prompt, don't have to relogin")
		return False
	else:
		debug("it is a login prompt, you need to re-login")
		return True
	 

def switch_find_shell_prompt(tn):
	TIMEOUT = 10
	tn.write(('' + '\n').encode('ascii'))
	time.sleep(1)

	tn.write(('\x03').encode('ascii'))
	time.sleep(2)
	tn.write(('\x03').encode('ascii'))
	time.sleep(2)
	tn.write(('\x03').encode('ascii'))
	time.sleep(2)
	debug("See what prompt the console is at")
	output = tn.expect([re.compile(b"#")],timeout=TIMEOUT)
	debug(output)
	if output[0] < 0: 
		debug("It is a NOT login prompt")
		return False
	else:
		debug("it is a login prompt")
		return True

def dict_lacp_boot_update(**kwargs):
    #dict_lacp_boot_update(dir_list=stat_dir_list,dut="dut1",mem=8,result=boot_result)
	tkwargs = {}
	for key, value in kwargs.items():
		tkwargs[key]=value
	debug(tkwargs)
	statlist = tkwargs["dir_list"] 
	dut = tkwargs["dut_name"] 
	mem = tkwargs["mem"]
	bootstats = tkwargs["result"]
	test = 0
	debug("!!!!!! print boot testing statistics here")
	debug(bootstats)
	for blist in bootstats:
		test+=1
		for b in blist:
			for lacp in statlist:
				if lacp["member"] == mem and lacp["test"] == test:
					working_lacp = lacp
					break
			reason = b["reason"]
			if b["tx_port"] == "1/1/2":
				host = "host2"
			elif b["tx_port"] == "1/1/1":
				host = "host1"
			else:
				ErrorHandler('Error identifying TX port', b)
			pkt_loss = b['loss_pkts'] 
			loss_time = b["loss_time"]
			try:
				tier = b["tier"]
			except Exception as e:
				tier = 0

			if tier == 1 and reason=="1st-down" and host=="host1":
				working_lacp["E4"] = pkt_loss
				working_lacp["F4"] = loss_time
			if tier == 1 and reason=="1st-down" and host=="host2":
				working_lacp["E5"] = pkt_loss
				working_lacp["F5"] = loss_time
			if tier == 1 and reason=="2nd-down" and host=="host1":
				working_lacp["E7"] = pkt_loss
				working_lacp["F7"] = loss_time
			if tier == 1 and reason=="2nd-down" and host=="host2":
				working_lacp["E8"] = pkt_loss
				working_lacp["F8"] = loss_time
			if tier == 1 and reason=="2nd-up" and host=="host1":
				working_lacp["E10"] = pkt_loss
				working_lacp["F10"] = loss_time
			if tier == 1 and reason=="2nd-up" and host=="host2":
				working_lacp["E11"] = pkt_loss
				working_lacp["F11"] = loss_time
			if tier == 1 and reason=="1st-up" and host=="host1":
				working_lacp["E13"] = pkt_loss
				working_lacp["F13"] = loss_time
			if tier == 1 and reason=="1st-up" and host=="host2":
				working_lacp["E14"] = pkt_loss
				working_lacp["F14"] = loss_time

			if tier == 2 and reason=="1st-down" and host=="host1":
				working_lacp["E16"] = pkt_loss
				working_lacp["F16"] = loss_time
			if tier == 2 and reason=="1st-down" and host=="host2":
				working_lacp["E17"] = pkt_loss
				working_lacp["F17"] = loss_time
			if tier == 2 and reason=="2nd-down" and host=="host1":
				working_lacp["E19"] = pkt_loss
				working_lacp["F19"] = loss_time
			if tier == 2 and reason=="2nd-down" and host=="host2":
				working_lacp["E20"] = pkt_loss
				working_lacp["F20"] = loss_time
			if tier == 2 and reason=="2nd-up" and host=="host1":
				working_lacp["E22"] = pkt_loss
				working_lacp["F22"] = loss_time
			if tier == 2 and reason=="2nd-up" and host=="host2":
				working_lacp["E23"] = pkt_loss
				working_lacp["F23"] = loss_time
			if tier == 2 and reason=="1st-up" and host=="host1":
				working_lacp["E25"] = pkt_loss
				working_lacp["F25"] = loss_time
			if tier == 2 and reason=="1st-up" and host=="host2":
				working_lacp["E26"] = pkt_loss
				working_lacp["F26"] = loss_time

			if dut=="dut1" and host == "host1" and reason == "down":
				working_lacp["E28"] = pkt_loss
				working_lacp["F28"] = loss_time
			elif dut=="dut1" and host == "host2" and reason == "down":
				working_lacp["E29"] = pkt_loss
				working_lacp["F29"] = loss_time
			elif dut=="dut2" and host == "host1" and reason == "down":
				working_lacp["E30"] = pkt_loss
				working_lacp["F30"] = loss_time
			elif dut=="dut2" and host == "host2" and reason == "down":
				working_lacp["E31"] = pkt_loss	
				working_lacp["F31"] = loss_time

			elif dut=="dut1" and host == "host1" and reason == "up":
				working_lacp["E33"] = pkt_loss
				working_lacp["F33"] = loss_time
			elif dut=="dut1" and host == "host2" and reason == "up":
				working_lacp["E34"] = pkt_loss	
				working_lacp["F34"] = loss_time	
			elif dut=="dut2" and host == "host1" and reason == "up":
				working_lacp["E35"] = pkt_loss
				working_lacp["F35"] = loss_time
			elif dut=="dut2" and host == "host2" and reason == "up":
				working_lacp["E36"] = pkt_loss
				working_lacp["F36"] = loss_time

			elif dut=="dut3" and host == "host1" and reason == "down":
				working_lacp["E38"] = pkt_loss
				working_lacp["F38"] = loss_time
			elif dut=="dut3" and host == "host2" and reason == "down":
				working_lacp["E39"] = pkt_loss
				working_lacp["F39"] = loss_time
			elif dut=="dut4" and host == "host1" and reason == "down":
				working_lacp["E40"] = pkt_loss
				working_lacp["F40"] = loss_time
			elif dut=="dut4" and host == "host2" and reason == "down":
				working_lacp["E41"] = pkt_loss
				working_lacp["F41"] = loss_time

			elif dut=="dut3" and host == "host1" and reason == "up":
				working_lacp["E43"] = pkt_loss
				working_lacp["F43"] = loss_time
			elif dut=="dut3" and host == "host2" and reason == "up":
				working_lacp["E44"] = pkt_loss
				working_lacp["F44"] = loss_time
			elif dut=="dut4" and host == "host1" and reason == "up":
				working_lacp["E45"] = pkt_loss
				working_lacp["F45"] = loss_time
			elif dut=="dut4" and host == "host2" and reason == "up":
				working_lacp["E46"] = pkt_loss
				working_lacp["F46"] = loss_time

def create_excel_sheets(stat_dir_list,image,**kwargs):
	runtime = kwargs["runtime"]
	mem = kwargs["mem"]
	for i in range(1,runtime+1):
		d = dict_lacp_blank(mem,image)
		d["test"] = i
		d['B4'] = "Test {}".format(i)
		d["sheetname"] = "LACP-{} Test{}".format(mem,i)
		stat_dir_list.append(d)

def dict_lacp_blank(member,image,**kwargs):
	if "runtime" in kwargs:
		runtime = kwargs["runtime"]
	else:
		runtime = 2
	lacp = {}
	lacp['member'] = member
	if member ==2: 
		lacp['B1:F1'] = "2-member LACP trunk"
	else:
		lacp['B1:F1'] = "8-member LACP trunk"
	lacp['B2:F2'] = image
	lacp['B3'] = "No."
	lacp['C3'] = "Action"
	lacp['D3:E3'] = "Frame Loss"
	lacp['F3'] = "Loss Time Sec"
	lacp['B4'] = "Test #"

	lacp['C4'] = "Tier-1:un-plug 1st active link"
	lacp['D4'] = "host1"
	lacp['D5'] = "host2"

	if member == 2:
		lacp['C7'] = "Tier-1: un-plug 2nd active link"
	elif member == 8:
		lacp['C7'] = "Tier-1: un-plug 3~4 active link"
	lacp['D7'] = "host1"
	lacp['D8'] = "host2"

	if member == 2:
		lacp['C10'] = "Tier-1: re-connect 2nd link"
	elif member == 8:
		lacp['C10'] = "Tier-1: re-connect 3~4 link"
	lacp['D10'] = "host1"
	lacp['D11'] = "host2"
	
	lacp['C13'] = "Tier-1: re-connect 1st active link"
	lacp['D13'] = "host1"
	lacp['D14'] = "host2"

	##########Tier 2 unplug rows
	lacp['C16'] = "Tier-2:un-plug 1st active link"
	lacp['D16'] = "host1"
	lacp['D17'] = "host2"

	if member == 2:
		lacp['C19'] = "Tier-2: un-plug 2nd active link"
	elif member == 8:
		lacp['C19'] = "Tier-2: un-plug 3~4 active link"
	lacp['D19'] = "host1"
	lacp['D20'] = "host2"

	if member == 2:
		lacp['C22'] = "Tier-2: re-connect 2nd link"
	elif member == 8:
		lacp['C22'] = "Tier-2: re-connect 3~4 link"
	lacp['D22'] = "host1"
	lacp['D23'] = "host2"
	
	lacp['C25'] = "Tier-2: re-connect 1st active link"
	lacp['D25'] = "host1"
	lacp['D26'] = "host2"


	# boot rows
	lacp['C28'] = "Tier-1 DUT1 Down"
	lacp['D28'] = "host1"
	lacp['D29'] = "host2"
	lacp['C30'] = "Tier-1 DUT2 Down"
	lacp['D30'] = "host1"
	lacp['D31'] = "host2"

	lacp['C33'] = "Tier-1 DUT1 Up"
	lacp['D33'] = "host1"
	lacp['D34'] = "host2"
	lacp['C35'] = "Tier-1 DUT2 Up"
	lacp['D35'] = "host1"
	lacp['D36'] = "host2"

	lacp['C38'] = "Tier-2 DUT3 Down"
	lacp['D38'] = "host1"
	lacp['D39'] = "host2"
	lacp['C40'] = "Tier-2 DUT4 Down"
	lacp['D40'] = "host1"
	lacp['D41'] = "host2"

	lacp['C43'] = "Tier-2 DUT3 Up"
	lacp['D43'] = "host1"
	lacp['D44'] = "host2"
	lacp['C45'] = "Tier-2 DUT4 Up"
	lacp['D45'] = "host1"
	lacp['D46'] = "host2"

	return lacp

def dict_2_excel(dict_list,filename):
	filePath = filename
	if os.path.exists(filePath):
	    os.remove(filePath)
	else:
	    tprint("Can not delete the file as it doesn't exists: {}".format(filePath))

	workbook = xlsxwriter.Workbook(filename)
	highlight_format = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center',
    'valign': 'vcenter',
    'fg_color': 'yellow'})

	merge_format_table = workbook.add_format({
	'bold': 0,
	'border': 1,
	'align': 'center',
	'valign': 'vcenter'})

	merge_format_title = workbook.add_format({
	'bold': 1,
	'border': 1,
	'align': 'left',
	'valign': 'vcenter'})

	format_title = workbook.add_format({
	'bold': 1,
	'border': 1,
	'align': 'left',
	'valign': 'vcenter'})

	full_border = workbook.add_format({"border":1})
	
	for stat in dict_list:
		sheetname = stat["sheetname"]
		stat.pop("sheetname")
		stat.pop("member")
		stat.pop("test")
		worksheet = workbook.add_worksheet(sheetname)
		worksheet.set_column("C:C", 30)
		worksheet.set_column("D:E", 15)
		worksheet.set_column("F:F", 25)
		worksheet.conditional_format('B4:F46', {'type':'blanks', 'format': full_border})
		for key, value in stat.items():
			if ":" in key: 
				worksheet.merge_range(key,value,merge_format_title)
			elif key=="C3" or key=="B3" or key=="F3":
				worksheet.write(key,value,format_title)
			else:
				worksheet.write(key,value,full_border)
	
	workbook.close()		 

def exel_2_member_lacp_blank(filename,sheetname,title,build,test_num):

	workbook = xlsxwriter.Workbook(filename)
	worksheet = workbook.add_worksheet(sheetname)
	#worksheet.write(0,0,"6.2.0 Interim Build 168")
	#worksheet.conditional_format( 'B4:F26' , { 'type' : 'no_blanks' , 'format' : border_format} )

	worksheet.set_landscape()
	worksheet.set_paper(8)
	worksheet.set_margins(0.787402, 0.787402, 0.5, 0.787402)

	apply_border_to_range(
		workbook,
		worksheet,
		{
			"range_string": "B2:F26",
			"border_style": 5,
		},
	)

	highlight_format = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center',
    'valign': 'vcenter',
    'fg_color': 'yellow'})

	merge_format_table = workbook.add_format({
	'bold': 0,
	'border': 1,
	'align': 'center',
	'valign': 'vcenter'})

	merge_format_title = workbook.add_format({
	'bold': 1,
	'border': 1,
	'align': 'left',
	'valign': 'vcenter'})

	full_border = workbook.add_format({"border":1})
	
	# for i in range(3,27):
	# 	for j in range(1,6):
	# 		worksheet.write(i,j,blank,full_border)
			
	worksheet.merge_range('B1:H1',title,merge_format_title)
	worksheet.merge_range('B2:F2',build,merge_format_table)
	worksheet.merge_range('D3:E3',"Frame Loss",merge_format_table)
	worksheet.conditional_format('B4:F26', {'type':'blanks', 'format': full_border})
	worksheet.write(2,1,"No.",full_border)
	worksheet.write(2,2,"Action",full_border)
	
	worksheet.write(2,5,"Loss Time Sec",full_border )
	worksheet.write(3,1,test_num,full_border)
	worksheet.write(3,2,"un-plug 1st active link",full_border)
	worksheet.write(3,3,"host1",full_border)
	worksheet.write(4,3,"host2",full_border)

	worksheet.write(6,2,"un-plug 2nd active link",full_border)
	worksheet.write(6,3,"host1",full_border)
	worksheet.write(7,3,"host2",full_border)

	worksheet.write(9,2,"re-connect 2nd link",full_border)
	worksheet.write(9,3,"host1",full_border)
	worksheet.write(10,3,"host2",full_border)

	worksheet.write(12,2,"re-connect 1st link",full_border)
	worksheet.write(12,3,"host1",full_border)
	worksheet.write(13,3,"host2",full_border)

	worksheet.write(15,2,"Tier-1 DUT2 Down",full_border)
	worksheet.write(15,3,"host1",full_border)
	worksheet.write(16,2,"Tier-1 DUT1  Down",full_border)
	worksheet.write(16,3,"host2",full_border)

	worksheet.write(18,2,"Tier-1 DUT2 Up",full_border)
	worksheet.write(18,3,"host1",full_border)
	worksheet.write(19,2,"Tier-1 DUT1  Up",full_border)
	worksheet.write(19,3,"host2",full_border)

	worksheet.write(21,2,"Tier-2 DUT3 Down",full_border)
	worksheet.write(21,3,"host1",full_border)
	worksheet.write(22,2,"Tier-2 DUT4 Down",full_border)
	worksheet.write(22,3,"host2",full_border)

	worksheet.write(24,2,"Tier-2 DUT3 Up",full_border)
	worksheet.write(24,3,"host1",full_border)
	worksheet.write(25,2,"Tier-2 DUT4 Up",full_border)
	worksheet.write(25,3,"host2",full_border)

	workbook.close()

def exel_8_member_lacp_blank(filename,sheetname,title,build,test_num):

	workbook = xlsxwriter.Workbook(filename)
	worksheet = workbook.add_worksheet(sheetname)
	#worksheet.write(0,0,"6.2.0 Interim Build 168")
	#worksheet.conditional_format( 'B4:F26' , { 'type' : 'no_blanks' , 'format' : border_format} )

	worksheet.set_landscape()
	worksheet.set_paper(8)
	worksheet.set_margins(0.787402, 0.787402, 0.5, 0.787402)

	apply_border_to_range(
		workbook,
		worksheet,
		{
			"range_string": "B2:F26",
			"border_style": 5,
		},
	)

	highlight_format = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center',
    'valign': 'vcenter',
    'fg_color': 'yellow'})

	merge_format_table = workbook.add_format({
	'bold': 0,
	'border': 1,
	'align': 'center',
	'valign': 'vcenter'})

	merge_format_title = workbook.add_format({
	'bold': 1,
	'border': 1,
	'align': 'left',
	'valign': 'vcenter'})

	full_border = workbook.add_format({"border":1})
	
	# for i in range(3,27):
	# 	for j in range(1,6):
	# 		worksheet.write(i,j,blank,full_border)
			
	worksheet.merge_range('B1:H1',title,merge_format_title)
	worksheet.merge_range('B2:F2',build,merge_format_table)
	worksheet.merge_range('D3:E3',"Frame Loss",merge_format_table)
	worksheet.conditional_format('B4:F26', {'type':'blanks', 'format': full_border})
	worksheet.write(2,1,"No.",full_border)
	worksheet.write(2,2,"Action",full_border)
	
	worksheet.write(2,5,"Loss Time Sec",full_border )
	worksheet.write(3,1,test_num,full_border)
	worksheet.write(3,2,"un-plug 1st active link",full_border)
	worksheet.write(3,3,"host1",full_border)
	worksheet.write(4,3,"host2",full_border)

	worksheet.write(6,2,"un-plug 3~4 active link",full_border)
	worksheet.write(6,3,"host1",full_border)
	worksheet.write(7,3,"host2",full_border)

	worksheet.write(9,2,"re-connect 3~4 link",full_border)
	worksheet.write(9,3,"host1",full_border)
	worksheet.write(10,3,"host2",full_border)

	worksheet.write(12,2,"re-connect 1st link",full_border)
	worksheet.write(12,3,"host1",full_border)
	worksheet.write(13,3,"host2",full_border)

	worksheet.write(15,2,"Tier-1 DUT2 Down",full_border)
	worksheet.write(15,3,"host1",full_border)
	worksheet.write(16,2,"Tier-1 DUT1  Down",full_border)
	worksheet.write(16,3,"host2",full_border)

	worksheet.write(18,2,"Tier-1 DUT2 Up",full_border)
	worksheet.write(18,3,"host1",full_border)
	worksheet.write(19,2,"Tier-1 DUT1  Up",full_border)
	worksheet.write(19,3,"host2",full_border)

	worksheet.write(21,2,"Tier-2 DUT3 Down",full_border)
	worksheet.write(21,3,"host1",full_border)
	worksheet.write(22,2,"Tier-2 DUT4 Down",full_border)
	worksheet.write(22,3,"host2",full_border)

	worksheet.write(24,2,"Tier-2 DUT3 Up",full_border)
	worksheet.write(24,3,"host1",full_border)
	worksheet.write(25,2,"Tier-2 DUT4 Up",full_border)
	worksheet.write(25,3,"host2",full_border)
	workbook.close()

def parse_linerate(result):
	linerate = {}
	num = 0
	linerate_list = []
	port_traffic_list = []
	for line in result:
		if "port" in line:
			line_list = line.strip().split("|")
			#print(line_list)
			port_traffic = [i.strip() for i in line_list if i !=""]
			debug(port_traffic)
			#if port_traffic[0] not in linerate:
			linerate[port_traffic[0]] = {}
			linerate[port_traffic[0]]["TX"] = port_traffic[1]
			linerate[port_traffic[0]]["TX_Rate"] = float(port_traffic[2].strip("Mbps"))
			linerate[port_traffic[0]]["RX"] = port_traffic[3]
			linerate[port_traffic[0]]["RX_Rate"] = float(port_traffic[4].strip("Mbps"))
			 
	debug(linerate)
	return linerate 
		 
	 
def parse_mclag_list(result):
	mclag = {}
	found = 0
	for line in result:
		if "-----" not in line:
			temp = line
			if found == 1: 
				line_list = line.strip().split("     ")
				if len(line_list) >= 2:
					key = line_list[0]
					value = line_list[-1]
					#print(line_list)
					mclag[name][key.strip()] = value.strip()
		else:
			name = temp.strip()
			mclag[name] = {}
			found = 1
	return mclag

def find_active_trunk_port(dut):
	result = collect_show_cmd(dut,"diag switch mclag list")
	mclag = parse_mclag_list(result)
	for k, v in mclag.items():
		if "core" in k or "FlInK1_MLAG" in k:
			if "-" in mclag[k]['Local ports']:
				start,finish = mclag[k]['Local ports'].split('-')
				start = int(start)
				finish = int(finish)
				core_ports = ["port"+ str(i) for i in range(start,finish+1)]
			else:
				core_ports = mclag[k]['Local ports'].split(',')
				try:
					debug("core_ports after split format like 2,4: {}".format(core_ports))
				except Exception as e:
					debug("find_active_trunk_port: not able to print valuable: core_ports")
				core_ports = ["port"+ p for p in core_ports]
	debug(mclag)
	debug(core_ports)
	 
	result = collect_show_cmd(dut,"diag switch physical linerate up")
	send_ctrl_c_cmd(dut)
	debug(result) 
	#print(result)
	TEST_RATE = 300
	linerate = parse_linerate(result)
	for p in core_ports:
		if p in linerate:
			if linerate[p]["TX_Rate"] > TEST_RATE or linerate[p]["RX_Rate"] > TEST_RATE:
				Info("On this switch, {} is active".format(p))
				return p
	return None

def loop_command_output(dut,cmd,**kwargs):
	if "timeout" in kwargs:
		timeout = kwargs['timeout']
	else:
		timeout = 3
	result = collect_show_cmd(dut,cmd,t=timeout)
	#print_show_cmd(dut,cmd,t=timeout)
	send_ctrl_c_cmd(dut)
	return (result)

def seperate_ip_mask(ip_addr):
	if ":" in ip_addr:
		regex = r'([0-9a-fA-F:]+)\/([0-9]+)'
		matched = re.search(regex,ip_addr)
		if matched:
			ip = matched.group(1)
			net = matched.group(2)
			return ip,net
		return None,None
	else:
		regex = r'([0-9]+\.[0-9]+\.[0-9]+\.[0-9]+)\/([0-9]+)'
		regex = r'([0-9.]+[0-9]+)\/([0-9]+)'
		matched = re.search(regex,ip_addr)
		if matched:
			ip = matched.group(1)
			net = matched.group(2)
			return ip,net
		return None,None

def find_inactive_trunk_port(dut):
	# print("========================= find_inactive_trunk port")
	result = collect_show_cmd(dut,"diag switch mclag list")
	mclag = parse_mclag_list(result)
	# print(mclag)
	for k, v in mclag.items():
		if "core" in k or "FlInK1_MLAG" in k:
			if "-" in mclag[k]['Local ports']:
				start,finish = mclag[k]['Local ports'].split('-')
				start = int(start)
				finish = int(finish)
				core_ports = ["port"+ str(i) for i in range(start,finish+1)]
			else:
				core_ports = mclag[k]['Local ports'].split(',')
				debug("core_ports after split ,".format(core_ports))
				core_ports = ["port"+ p for p in core_ports]
	debug(mclag)
	debug(core_ports)
	 
	result = collect_show_cmd(dut,"diag switch physical linerate up")
	send_ctrl_c_cmd(dut)
	# for line in result:
	# 	print(line)
	 
	TEST_RATE = 20
	linerate = parse_linerate(result)
	inactive_ports = []
	for p in core_ports:
		if p in linerate:
			if linerate[p]["TX_Rate"] < TEST_RATE and linerate[p]["RX_Rate"] < TEST_RATE:
				Info("On this switch, {} is INactive".format(p))
				inactive_ports.append(p)

	return inactive_ports

	
	#mclag_list = parse_mclag_list(result)
 


def eprint(*args, **kwargs):
    print(*args, file=sys.stderr, **kwargs)


def print_interactive_line():
	print("-------------------------------- Go to the lab to take action --------------------")

def print_dash_line():
	print("------------------------------------------------------------------------------------------------------")

def print_double_line():
	print("======================================================================================================")

def print_hash_line(msg=""):
	print(f"\n######################################{msg}#############################################")

def find_subnet(ipaddr,mask_length):
	import ipaddress
	n = f"{ipaddr}/{mask_length}"
	net = str(ipaddress.ip_network(n, strict=False))
	subnet = (net.split('/'))[0]
	print(subnet)
	return subnet

def print_test_subject(testcase,description):
	print(f"======================= Testcase #{testcase}: {description} =============================")

def smooth_cli_line(line):
	line = line.strip()
	debug(f'old line = {line}')
	if "More" in line:
		line = line.replace("--More--",'')
		line = line.replace("\r",'')
		line = line.strip()
	debug(f'new line = {line}')
	return line

def generate_ip_addresses_24(start_ip, count):
	ip_list = []
	ip_parts = start_ip.split('.')  # Split the start IP into octets

	for _ in range(count):
		ip_list.append(".".join(ip_parts))
		ip_parts[2] = str(int(ip_parts[2]) + 1)  # Increment the third octet
		if int(ip_parts[2]) > 255:
			ip_parts[2] = '0'

	return ip_list

def list_routers_rancid(folder):
  """
  Lists device configuration filenames from a RANCID-style folder using os module.

  Args:
    folder_name: The path (string) to the RANCID group folder.

  Returns:
    A list of filenames (strings) representing the devices found.
    Returns an empty list on error.
  """
  folder_name = f'/home/build/ops/network/rancid/gfiber/{folder}'
  device_list = []
  try:
    # Check if the path exists and is a directory
    if not os.path.isdir(folder_name):
      print(f"Error: Path '{folder_name}' is not a valid directory.", file=sys.stderr)
      return []

    # List comprehension to get entries and filter for files in one step
    device_list = [
        entry for entry in os.listdir(folder_name)
        if os.path.isfile(os.path.join(folder_name, entry))
    ]
  except FileNotFoundError:
    print(f"Error: Folder not found at '{folder_name}'", file=sys.stderr)
    return []
  except OSError as e:
    print(f"Error accessing folder '{folder_name}': {e}", file=sys.stderr)
    return []

  return device_list

def find_files(pattern):
	# List all items in the current directory
	files_found = []
	for filename in os.listdir('.'):
	    # Check if the item is a file and ends with '.yaml'
	    if os.path.isfile(filename) and filename.endswith(pattern):
	        dprint(filename)
	        files_found.append(filename)
	return files_found

def build_rancid_yaml(folder,router_types,log_folder):
	# Define the directory you want to scan
	#my own rancid
	#directory = '/google/src/cloud/mikezh/mikezh/ops/network/rancid/gfiber/juniper'
	#shared rancid
	directory = f'/home/build/ops/network/rancid/gfiber/{folder}'

	file_names=[]
	# Walk through the directory
	for root, dirs, files in os.walk(directory):
	    for file in files:
	        file_names.append(file)

	# Example values for the YAML structure
	#log_folder = "Log_data"

	commands_list = [
	    "show version",
	    "show interfaces descriptions",
	    "show interfaces terse",
	    "show isis adjacency",
	    #"set cli screen-length 0",
	    "show interfaces brief",
	    "show chassis hardware"
	]

	global_login = "mikezh"


	# Construct the dictionary structure for the YAML

 	#router_types=["cr","pr","dr"]
	hostname_list = []
	for router_type in router_types:
	    for name in file_names:
	        #if "cr" in name or "dr" in name or "pr" in name:
	        if name.startswith(router_type):
	            print(f"found device name:{name}")
	            hostname_list.append(name)
	    # Populate the device_list with empty mgmt_ip and model fields
	device_list = [{"hostname": hostname, "mgmt_ip": "", "model": ""} for hostname in hostname_list]
	yaml_structure = {
	    "Log_folder": log_folder,
	    "Commands_list": commands_list,
	    "Global_login": global_login,
	    "Device_list": device_list
	}

	rancid_folder=f'{folder}_{"_".join(router_types)}'
	# Define the output YAML file path
	output_file = f"rancid_{rancid_folder}.yaml"

	# Write the dictionary to a YAML file
	with open(output_file, 'w') as file:
	    yaml.dump(yaml_structure, file, default_flow_style=False)

	print(f"YAML file '{output_file}' has been created with the specified format.")
	return output_file

def build_rancid_yaml_v2(folder,router_types,log_folder,output_file):
	# Define the directory you want to scan
	#directory = '/google/src/cloud/mikezh/mikezh/ops/network/rancid/gfiber/juniper'
	directory = f'/home/build/ops/network/rancid/gfiber/{folder}'

	file_names=[]
	# Walk through the directory
	for root, dirs, files in os.walk(directory):
	    for file in files:
	        file_names.append(file)

	# Example values for the YAML structure
	#log_folder = "Log_data"

	commands_list = [
	    "show version",
	    "show interfaces descriptions",
	    "show interfaces terse",
	    "show isis adjacency",
	    #"set cli screen-length 0",
	    "show interfaces brief",
	    "show chassis hardware"
	]

	global_login = "mikezh"


	# Construct the dictionary structure for the YAML

 	#router_types=["cr","pr","dr"]
	hostname_list = []
	for router_type in router_types:
	    for name in file_names:
	        #if "cr" in name or "dr" in name or "pr" in name:
	        if name.startswith(router_type):
	            print(name)
	            hostname_list.append(name)
	    # Populate the device_list with empty mgmt_ip and model fields
	print(hostname_list)
	device_list = [{"hostname": hostname, "mgmt_ip": "", "model": ""} for hostname in hostname_list]
	yaml_structure = {
	    "Log_folder": log_folder,
	    "Commands_list": commands_list,
	    "Global_login": global_login,
	    "Device_list": device_list
	}

	# rancid_folder=f'{folder}_{"_".join(router_types)}'
	# # Define the output YAML file path
	# output_file = f"rancid_{rancid_folder}.yaml"

	# Write the dictionary to a YAML file
	with open(output_file, 'w') as file:
	    yaml.dump(yaml_structure, file, default_flow_style=False)

	print(f"YAML file '{output_file}' has been created with the specified format.")
	# return output_file


if __name__ == "__main__":
# 	debug("test debug")
    #tn = telnet_switch("10.132.10.142", "5022")
    #tn = telnet_switch("10.132.10.142", "5024")
    #tn = telnet_switch("10.132.10.142", "5017")
    tn = telnet_switch("10.92.44.6", "6010")
    tn = telnet_switch("10.92.44.6", "6011")
    tn = telnet_switch("10.92.44.5", "6005")
    exit()
    